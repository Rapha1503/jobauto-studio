from __future__ import annotations

import json
import os
import re
import tempfile
from datetime import date
from pathlib import Path
from typing import Literal, Protocol
from uuid import uuid4

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field

from jobauto.application_service import RunApplicationService, RunRequest
from jobauto.candidate_snapshot import CandidateProfileRepository, CandidateSnapshot
from jobauto.excel_schema import CANDIDATE_ID_COLUMN, ensure_tracker_schema, header_map
from jobauto.offer_catalog import (
    MIN_FULL_DESCRIPTION_LENGTH,
    OfferCandidate,
    SelectedOffer,
    append_selected_offers,
    canonical_url,
    existing_urls_from_workbook,
    load_candidates,
    parse_posted_date,
)
from jobauto.run_store import utc_now
from jobauto.search_preferences import SearchEvaluation, SearchOffer
from jobauto.tracker_io import save_workbook_atomically, tracker_lock

CampaignStatus = Literal[
    "ready",
    "running",
    "cancelling",
    "cancelled",
    "completed",
    "partial",
    "failed",
]
CampaignDecision = Literal["selected", "duplicate", "rejected", "not_selected"]


class ApplicationService(Protocol):
    def start(self, request: RunRequest) -> str: ...

    def execute(self, run_id: str): ...

    def get(self, run_id: str): ...


class AvailabilityCheck(Protocol):
    status: str
    reason: str


class OfferAvailabilityVerifier(Protocol):
    def verify(self, url: str) -> AvailabilityCheck: ...


class StudioCampaignItem(BaseModel):
    model_config = ConfigDict(extra="forbid")

    offer: OfferCandidate
    canonical_url: str
    decision: CampaignDecision
    evaluation: SearchEvaluation | None = None
    reasons: list[str] = Field(default_factory=list)
    excel_row: int | None = Field(default=None, ge=2)
    run_id: str | None = None
    run_status: str | None = None
    run_phase: str | None = None
    run_blockers: list[str] = Field(default_factory=list)
    agent_call_count: int = Field(default=0, ge=0)
    repair_call_count: int = Field(default=0, ge=0)
    agent_latency_ms: int = Field(default=0, ge=0)
    agent_token_estimate: int = Field(default=0, ge=0)
    latest_agent_phase: str | None = None
    latest_agent_status: str | None = None
    tracker_artifacts_synced: bool = False
    tracker_sync_error: str | None = None


class StudioCampaignRecord(BaseModel):
    model_config = ConfigDict(extra="forbid")

    campaign_id: str = Field(min_length=8, max_length=160)
    candidate_id: str = Field(min_length=2, max_length=80)
    profile_path: Path
    tracker_path: Path
    status: CampaignStatus
    created_at: str
    updated_at: str
    requested_limit: int = Field(ge=1, le=100)
    processing_limit: int | None = Field(default=None, ge=1, le=100)
    campaign_dir: Path
    items: list[StudioCampaignItem] = Field(default_factory=list)
    cancel_requested_at: str | None = None
    cancelled_at: str | None = None

    @property
    def selected_count(self) -> int:
        return sum(item.decision == "selected" for item in self.items)

    @property
    def reserve_count(self) -> int:
        return sum(item.decision == "not_selected" for item in self.items)

    @property
    def effective_limit(self) -> int:
        return self.processing_limit or self.requested_limit


class StudioCampaignStore:
    def __init__(self, root: Path) -> None:
        self.root = root.expanduser().resolve()
        self.root.mkdir(parents=True, exist_ok=True)

    def create(self, record: StudioCampaignRecord) -> StudioCampaignRecord:
        record.campaign_dir.mkdir(parents=True, exist_ok=False)
        return self.save(record)

    def save(self, record: StudioCampaignRecord) -> StudioCampaignRecord:
        if not record.campaign_dir.is_dir():
            raise FileNotFoundError(f"campaign directory does not exist: {record.campaign_dir}")
        _atomic_write_json(
            record.campaign_dir / "campaign.json",
            record.model_dump(mode="json"),
        )
        return record

    def get(self, campaign_id: str) -> StudioCampaignRecord:
        path = self.root / campaign_id / "campaign.json"
        if not path.is_file():
            raise FileNotFoundError(f"Studio campaign not found: {campaign_id}")
        return StudioCampaignRecord.model_validate_json(path.read_text(encoding="utf-8"))

    def list_for_candidate(
        self,
        candidate_id: str,
        *,
        limit: int = 10,
    ) -> list[StudioCampaignRecord]:
        records: list[StudioCampaignRecord] = []
        for path in self.root.glob("*/campaign.json"):
            try:
                record = StudioCampaignRecord.model_validate_json(path.read_text(encoding="utf-8"))
            except (OSError, ValueError):
                continue
            if record.candidate_id == candidate_id:
                records.append(record)
        return sorted(records, key=lambda record: record.updated_at, reverse=True)[:limit]


class StudioCampaignService:
    def __init__(
        self,
        *,
        repository: CandidateProfileRepository,
        application_service: RunApplicationService | ApplicationService,
        store: StudioCampaignStore,
        availability_verifier: OfferAvailabilityVerifier | None = None,
    ) -> None:
        self.repository = repository
        self.application_service = application_service
        self.store = store
        self.availability_verifier = availability_verifier

    def create(
        self,
        *,
        profile_path: Path,
        tracker_path: Path,
        candidates: list[dict[str, object]],
        limit: int,
        today: date | None = None,
    ) -> StudioCampaignRecord:
        tracker = tracker_path.expanduser().resolve()
        if tracker.suffix.casefold() != ".xlsx" or not tracker.is_file():
            raise ValueError("tracker must be an existing .xlsx workbook")
        if not candidates:
            raise ValueError("campaign import contains no offers")

        snapshot = self.repository.load_snapshot(profile_path)
        processing_limit = min(
            limit,
            snapshot.submission_preferences.max_applications_per_campaign,
        )
        campaign_id = f"{snapshot.profile.candidate_id}-{uuid4().hex[:12]}"
        campaign_dir = self.store.root / campaign_id
        now = utc_now()
        record = StudioCampaignRecord(
            campaign_id=campaign_id,
            candidate_id=snapshot.profile.candidate_id,
            profile_path=profile_path.expanduser().resolve(),
            tracker_path=tracker,
            status="ready",
            created_at=now,
            updated_at=now,
            requested_limit=limit,
            processing_limit=processing_limit,
            campaign_dir=campaign_dir,
        )
        self.store.create(record)
        source_path = campaign_dir / "candidates.json"
        _atomic_write_json(source_path, {"offers": candidates})
        imported = load_candidates(source_path)
        if not imported:
            raise ValueError("campaign import contains no valid offers")

        with tracker_lock(tracker):
            items = self._select_and_activate_items(
                record=record,
                imported=imported,
                snapshot=snapshot,
                processing_limit=processing_limit,
                requested_limit=limit,
                today=today,
            )

        record = record.model_copy(update={"items": items, "updated_at": utc_now()})
        return self.store.save(record)

    def attach_completed_run(
        self,
        *,
        profile_path: Path,
        tracker_path: Path,
        offer: OfferCandidate,
        run_id: str,
        today: date | None = None,
    ) -> StudioCampaignRecord:
        """Attach a completed direct-offer run to the application queue without regenerating it."""
        tracker = tracker_path.expanduser().resolve()
        if tracker.suffix.casefold() != ".xlsx" or not tracker.is_file():
            raise ValueError("tracker must be an existing .xlsx workbook")
        snapshot = self.repository.load_snapshot(profile_path)
        run = self.application_service.get(run_id)
        if run.candidate_id != snapshot.profile.candidate_id:
            raise ValueError("run and profile belong to different candidates")
        if run.status != "completed":
            raise ValueError("only a completed run can enter the application queue")
        if not run.artifacts.get("cv") or not run.artifacts.get("letter"):
            raise ValueError("completed run has no approved CV and letter")
        for campaign in self.store.list_for_candidate(snapshot.profile.candidate_id, limit=100):
            if any(item.run_id == run_id for item in campaign.items):
                return campaign

        evaluation = snapshot.search_preferences.evaluate(
            _to_search_offer(offer, today=today or date.today()),
            today=today,
        )
        reasons = [finding.message for finding in evaluation.blockers]
        if reasons:
            reasons.insert(0, "Explicit offer retained by the candidate despite search filters.")
        item = StudioCampaignItem(
            offer=offer,
            canonical_url=canonical_url(offer.url),
            decision="selected",
            evaluation=evaluation,
            reasons=reasons,
            run_id=run_id,
            run_status=run.status,
            run_phase=run.current_phase,
            run_blockers=list(run.blockers),
        )
        (
            item.agent_call_count,
            item.repair_call_count,
            item.agent_latency_ms,
            item.agent_token_estimate,
            item.latest_agent_phase,
            item.latest_agent_status,
        ) = _run_observability(run)

        campaign_id = f"{snapshot.profile.candidate_id}-{uuid4().hex[:12]}"
        campaign_dir = self.store.root / campaign_id
        now = utc_now()
        record = StudioCampaignRecord(
            campaign_id=campaign_id,
            candidate_id=snapshot.profile.candidate_id,
            profile_path=profile_path.expanduser().resolve(),
            tracker_path=tracker,
            status="completed",
            created_at=now,
            updated_at=now,
            requested_limit=1,
            processing_limit=1,
            campaign_dir=campaign_dir,
            items=[item],
        )
        self.store.create(record)
        _atomic_write_json(campaign_dir / "candidates.json", {"offers": [offer.model_dump()]})

        item.excel_row = _find_tracker_row(
            tracker,
            candidate_id=record.candidate_id,
            offer_url=offer.url,
        )
        if item.excel_row is None:
            item.excel_row = append_selected_offers(
                tracker,
                [
                    SelectedOffer(
                        company=offer.company,
                        role=offer.role,
                        url=offer.url,
                        posted_at=offer.posted_at,
                        location=offer.location,
                        description=offer.description,
                        fit_score=evaluation.score,
                        source=offer.source,
                        experience_required=offer.experience_required,
                        contract_type=offer.contract_type,
                        salary_estimate=offer.salary_estimate,
                        semantic_fit_score=offer.semantic_fit_score,
                        semantic_fit_rationale=offer.semantic_fit_rationale,
                    )
                ],
                candidate_id=record.candidate_id,
            )[0]
        _sync_completed_artifacts(record, item, run.artifacts)
        return self.store.save(record.model_copy(update={"items": [item], "updated_at": utc_now()}))

    def _select_and_activate_items(
        self,
        *,
        record: StudioCampaignRecord,
        imported: list[OfferCandidate],
        snapshot: CandidateSnapshot,
        processing_limit: int,
        requested_limit: int,
        today: date | None,
    ) -> list[StudioCampaignItem]:
        existing = existing_urls_from_workbook(
            record.tracker_path,
            candidate_id=record.candidate_id,
        )
        seen: set[str] = set()
        items: list[StudioCampaignItem] = []
        eligible: list[tuple[int, StudioCampaignItem]] = []
        for order, candidate in enumerate(imported):
            url_key = canonical_url(candidate.url)
            if url_key in seen or url_key in existing:
                items.append(
                    StudioCampaignItem(
                        offer=candidate,
                        canonical_url=url_key,
                        decision="duplicate",
                        reasons=["Offer URL already exists in this import or tracker."],
                    )
                )
                continue
            seen.add(url_key)
            if len(candidate.description.strip()) < MIN_FULL_DESCRIPTION_LENGTH:
                items.append(
                    StudioCampaignItem(
                        offer=candidate,
                        canonical_url=url_key,
                        decision="rejected",
                        reasons=["Full offer text is required before ATS analysis."],
                    )
                )
                continue
            evaluation = snapshot.search_preferences.evaluate(
                _to_search_offer(candidate, today=today or date.today()),
                today=today,
            )
            item = StudioCampaignItem(
                offer=candidate,
                canonical_url=url_key,
                decision="selected" if evaluation.eligible else "rejected",
                evaluation=evaluation,
                reasons=[finding.message for finding in evaluation.blockers],
            )
            items.append(item)
            if evaluation.eligible:
                eligible.append((order, item))

        ranked = sorted(
            eligible,
            key=lambda entry: (*_campaign_rank_key(entry[1]), entry[0]),
        )
        selected_ids = {id(item) for _, item in ranked[:processing_limit]}
        for _, item in ranked[processing_limit:]:
            item.decision = "not_selected"
            if processing_limit < requested_limit:
                item.reasons = ["Eligible but outside the candidate's configured campaign limit."]
            else:
                item.reasons = ["Eligible but outside the requested campaign limit."]

        selected_items = [item for item in items if id(item) in selected_ids]
        self._activate_items(record, selected_items)
        return items

    def execute(self, campaign_id: str) -> StudioCampaignRecord:
        record = self.store.get(campaign_id)
        if record.cancel_requested_at is not None:
            return self._mark_cancelled(record)
        record = self._save_status(record, "running")
        while True:
            for item in record.items:
                latest = self.store.get(campaign_id)
                if latest.cancel_requested_at is not None:
                    return self._mark_cancelled(latest)
                if item.decision != "selected" or not item.run_id:
                    continue
                if item.run_status == "completed":
                    if not item.tracker_artifacts_synced and item.excel_row is not None:
                        try:
                            run = self.application_service.get(item.run_id)
                            _sync_completed_artifacts(record, item, run.artifacts)
                        except (OSError, KeyError, ValueError) as exc:
                            item.tracker_sync_error = f"{type(exc).__name__}: {exc}"
                        record = record.model_copy(
                            update={"items": record.items, "updated_at": utc_now()}
                        )
                        self.store.save(record)
                    continue
                if item.run_status in {"blocked", "failed"}:
                    continue
                try:
                    run = self.application_service.execute(item.run_id)
                    item.run_status = run.status
                    item.run_phase = run.current_phase
                    item.run_blockers = list(run.blockers)
                    if run.status == "completed" and item.excel_row is not None:
                        try:
                            _sync_completed_artifacts(record, item, run.artifacts)
                        except (OSError, KeyError, ValueError) as exc:
                            item.tracker_sync_error = f"{type(exc).__name__}: {exc}"
                except Exception as exc:
                    item.run_status = "failed"
                    item.run_phase = "execution_failed"
                    item.run_blockers = [f"{type(exc).__name__}: {exc}"]
                latest = self.store.get(campaign_id)
                record = latest.model_copy(update={"items": record.items, "updated_at": utc_now()})
                self.store.save(record)

            completed_count = sum(
                item.decision == "selected" and item.run_status == "completed"
                for item in record.items
            )
            missing = record.effective_limit - completed_count
            if missing <= 0:
                break
            replacements = sorted(
                (item for item in record.items if item.decision == "not_selected"),
                key=_campaign_rank_key,
            )[:missing]
            if not replacements:
                break
            for item in replacements:
                item.decision = "selected"
                item.reasons = ["Promoted after an earlier candidate did not produce documents."]
            self._activate_items(record, replacements)
            record = record.model_copy(update={"items": record.items, "updated_at": utc_now()})
            self.store.save(record)

        return self._save_status(record, _terminal_campaign_status(record))

    def request_cancel(self, campaign_id: str) -> StudioCampaignRecord:
        record = self.store.get(campaign_id)
        if record.status in {"completed", "cancelled"}:
            raise ValueError(f"campaign cannot be cancelled from status: {record.status}")
        now = utc_now()
        if record.status == "running":
            status: CampaignStatus = "cancelling"
            cancelled_at = None
        else:
            status = "cancelled"
            cancelled_at = now
        return self.store.save(
            record.model_copy(
                update={
                    "status": status,
                    "cancel_requested_at": now,
                    "cancelled_at": cancelled_at,
                    "updated_at": now,
                }
            )
        )

    def resume(self, campaign_id: str) -> StudioCampaignRecord:
        record = self.store.get(campaign_id)
        if record.status == "completed":
            return record
        if record.status not in {"cancelled", "partial", "failed"}:
            raise ValueError(f"campaign cannot be resumed from status: {record.status}")
        return self.store.save(
            record.model_copy(
                update={
                    "status": "ready",
                    "cancel_requested_at": None,
                    "cancelled_at": None,
                    "updated_at": utc_now(),
                }
            )
        )

    def expand(
        self,
        campaign_id: str,
        *,
        additional_count: int | None = None,
    ) -> StudioCampaignRecord:
        """Promote ranked eligible reserves without rerunning discovery."""
        record = self.store.get(campaign_id)
        snapshot = self.repository.load_snapshot(record.profile_path)
        configured_limit = snapshot.submission_preferences.max_applications_per_campaign
        available_capacity = configured_limit - record.selected_count
        if available_capacity <= 0:
            raise ValueError("The candidate's configured application limit is already reached.")
        reserves = sorted(
            (item for item in record.items if item.decision == "not_selected"),
            key=_campaign_rank_key,
        )
        if self.availability_verifier is not None:
            available_reserves: list[StudioCampaignItem] = []
            for item in reserves:
                check = self.availability_verifier.verify(str(item.offer.url))
                if check.status == "unavailable":
                    item.decision = "rejected"
                    item.reasons = [f"Offer unavailable before generation: {check.reason}"]
                else:
                    available_reserves.append(item)
            reserves = available_reserves
        if not reserves:
            self.store.save(
                record.model_copy(update={"items": record.items, "updated_at": utc_now()})
            )
            raise ValueError("This campaign has no eligible offers in reserve.")
        promote_count = min(
            additional_count if additional_count is not None else len(reserves),
            available_capacity,
            len(reserves),
        )
        if promote_count < 1:
            raise ValueError("additional_count must promote at least one offer")
        promoted = reserves[:promote_count]
        for item in promoted:
            item.decision = "selected"
            item.reasons = ["Promoted from the eligible reserve by the candidate."]
        with tracker_lock(record.tracker_path):
            self._activate_items(record, promoted)
        processing_limit = max(record.effective_limit, record.selected_count)
        updated = record.model_copy(
            update={
                "items": record.items,
                "processing_limit": processing_limit,
                "status": "ready",
                "updated_at": utc_now(),
            }
        )
        return self.store.save(updated)

    def get(self, campaign_id: str) -> StudioCampaignRecord:
        record = self.store.get(campaign_id)
        changed = False
        for item in record.items:
            if not item.run_id:
                continue
            try:
                run = self.application_service.get(item.run_id)
            except FileNotFoundError:
                continue
            observability = _run_observability(run)
            values = (
                run.status,
                run.current_phase,
                list(run.blockers),
                *observability,
            )
            current = (
                item.run_status,
                item.run_phase,
                item.run_blockers,
                item.agent_call_count,
                item.repair_call_count,
                item.agent_latency_ms,
                item.agent_token_estimate,
                item.latest_agent_phase,
                item.latest_agent_status,
            )
            if values != current:
                (
                    item.run_status,
                    item.run_phase,
                    item.run_blockers,
                    item.agent_call_count,
                    item.repair_call_count,
                    item.agent_latency_ms,
                    item.agent_token_estimate,
                    item.latest_agent_phase,
                    item.latest_agent_status,
                ) = values
                changed = True
        if changed:
            status = record.status
            if status == "running" and all(
                item.run_status not in {"pending", "running"}
                for item in record.items
                if item.decision == "selected"
            ):
                status = _terminal_campaign_status(record)
            record = record.model_copy(
                update={"items": record.items, "status": status, "updated_at": utc_now()}
            )
            self.store.save(record)
        return record

    def _save_status(
        self,
        record: StudioCampaignRecord,
        status: CampaignStatus,
    ) -> StudioCampaignRecord:
        updated = record.model_copy(update={"status": status, "updated_at": utc_now()})
        return self.store.save(updated)

    def _mark_cancelled(self, record: StudioCampaignRecord) -> StudioCampaignRecord:
        now = utc_now()
        return self.store.save(
            record.model_copy(
                update={
                    "status": "cancelled",
                    "cancelled_at": now,
                    "updated_at": now,
                }
            )
        )

    def _activate_items(
        self,
        record: StudioCampaignRecord,
        items: list[StudioCampaignItem],
    ) -> None:
        if not items:
            return
        rows = append_selected_offers(
            record.tracker_path,
            [
                SelectedOffer(
                    company=item.offer.company,
                    role=item.offer.role,
                    url=item.offer.url,
                    posted_at=item.offer.posted_at,
                    location=item.offer.location,
                    description=item.offer.description,
                    fit_score=item.evaluation.score,  # type: ignore[union-attr]
                    source=item.offer.source,
                    experience_required=item.offer.experience_required,
                    contract_type=item.offer.contract_type,
                    salary_estimate=item.offer.salary_estimate,
                    semantic_fit_score=item.offer.semantic_fit_score,
                    semantic_fit_rationale=item.offer.semantic_fit_rationale,
                )
                for item in items
            ],
            candidate_id=record.candidate_id,
        )
        for item, excel_row in zip(items, rows, strict=True):
            item.excel_row = excel_row
            try:
                item.run_id = self.application_service.start(
                    RunRequest(
                        profile_path=record.profile_path,
                        offer_text=item.offer.description,
                        offer_url=item.offer.url,
                        company=item.offer.company,
                        role=item.offer.role,
                    )
                )
                item.run_status = "pending"
                item.run_phase = "pending"
            except Exception as exc:
                item.run_status = "failed"
                item.run_phase = "start_failed"
                item.run_blockers = [f"{type(exc).__name__}: {exc}"]


def _campaign_rank_key(item: StudioCampaignItem) -> tuple[int, int]:
    deterministic_score = item.evaluation.score if item.evaluation else 0
    semantic_score = item.offer.semantic_fit_score
    return (
        -(semantic_score if semantic_score is not None else deterministic_score),
        -deterministic_score,
    )


def _run_observability(run) -> tuple[int, int, int, int, str | None, str | None]:
    terminal_by_call: dict[str, dict[str, object]] = {}
    latest_phase: str | None = None
    latest_status: str | None = None
    for index, event in enumerate(run.agent_events):
        phase = str(event.get("phase", "")).strip() or None
        status = str(event.get("status", "")).strip() or None
        if phase is not None:
            latest_phase = phase
            latest_status = status
        call_id = str(event.get("call_id") or f"event-{index}")
        if status in {"succeeded", "rejected", "failed"}:
            terminal_by_call[call_id] = event
    completed = list(terminal_by_call.values())
    repair_count = sum("repair" in str(event.get("phase", "")) for event in completed)
    latency_ms = sum(int(event.get("latency_ms") or 0) for event in completed)
    tokens = sum(int(event.get("total_tokens_estimate") or 0) for event in completed)
    return len(completed), repair_count, latency_ms, tokens, latest_phase, latest_status


def _to_search_offer(candidate: OfferCandidate, *, today: date) -> SearchOffer:
    salary, currency = _parse_salary(candidate.salary_estimate)
    experience_values = [
        value
        for value in (
            _parse_experience_years(candidate.experience_required),
            _parse_description_experience_years(candidate.description),
        )
        if value is not None
    ]
    return SearchOffer(
        company=candidate.company,
        title=candidate.role,
        source_url=candidate.url,
        description=candidate.description,
        location=candidate.location,
        contract=candidate.contract_type,
        experience_years=max(experience_values) if experience_values else None,
        posted_at=parse_posted_date(candidate, today=today),
        salary_annual=salary,
        salary_currency=currency,
    )


def _parse_experience_years(value: str | None) -> float | None:
    if not value:
        return None
    normalized = value.casefold()
    year_unit = r"(?:years?|yrs?|ans?|ann(?:e|é)es?)"
    values: list[float] = []
    for lower, upper in re.findall(
        rf"(\d+(?:[.,]\d+)?)\s*(?:-|–|—|to|à)\s*(\d+(?:[.,]\d+)?)\s*{year_unit}",
        normalized,
    ):
        values.extend((float(lower.replace(",", ".")), float(upper.replace(",", "."))))
    without_ranges = re.sub(
        rf"\d+(?:[.,]\d+)?\s*(?:-|–|—|to|à)\s*\d+(?:[.,]\d+)?\s*{year_unit}",
        "",
        normalized,
    )
    values.extend(
        float(number.replace(",", "."))
        for number in re.findall(rf"(\d+(?:[.,]\d+)?)\s*\+?\s*{year_unit}", without_ranges)
    )
    return max(values) if values else None


def _parse_description_experience_years(description: str) -> float | None:
    requirement_cues = re.compile(
        r"\b(?:applicant|candidate|profile|profil|you|your|vous|votre|must|required|"
        r"minimum|at least|au moins|requis|requise|exige|exigee|disposez|justifiez)\b"
    )
    experience_cues = re.compile(
        r"\b(?:experience|experiences|background|track record|years? of|ans? d['’e])"
    )
    values: list[float] = []
    for fragment in re.split(r"(?<=[.!?;])\s+|[\r\n]+", description.casefold()):
        if not requirement_cues.search(fragment) or not experience_cues.search(fragment):
            continue
        parsed = _parse_experience_years(fragment)
        if parsed is not None:
            values.append(parsed)
    return max(values) if values else None


def _parse_salary(value: str | None) -> tuple[int | None, str | None]:
    if not value:
        return None, None
    compact = value.casefold().replace(" ", "")
    amounts: list[int] = []
    for number, suffix in re.findall(r"(\d+(?:[.,]\d+)?)(k)?", compact):
        amount = float(number.replace(",", "."))
        if suffix:
            amount *= 1000
        if amount >= 10_000:
            amounts.append(round(amount))
    currency = "EUR" if any(token in compact for token in ("eur", "euro", "€")) else None
    return (min(amounts) if amounts else None), currency


def _terminal_campaign_status(record: StudioCampaignRecord) -> CampaignStatus:
    selected = [item for item in record.items if item.decision == "selected"]
    if not selected:
        return "completed"
    completed_count = sum(item.run_status == "completed" for item in selected)
    if completed_count >= record.effective_limit:
        return "completed"
    if all(item.run_status == "completed" for item in selected):
        return "completed"
    if any(item.run_status in {"completed", "blocked"} for item in selected):
        return "partial"
    return "failed"


def _sync_artifact_links(
    tracker_path: Path,
    excel_row: int,
    artifacts: dict[str, dict[str, object]],
    *,
    candidate_id: str,
    offer_url: str,
) -> None:
    paths: dict[str, Path] = {}
    for kind in ("cv", "letter"):
        path = Path(str(artifacts[kind]["pdf_path"])).expanduser().resolve()
        if not path.is_file():
            raise FileNotFoundError(f"completed {kind} PDF does not exist: {path}")
        paths[kind] = path

    with tracker_lock(tracker_path):
        workbook = load_workbook(tracker_path)
        try:
            sheet = (
                workbook["Postulations"]
                if "Postulations" in workbook.sheetnames
                else workbook.create_sheet("Postulations")
            )
            columns = ensure_tracker_schema(sheet)
            _claim_and_validate_tracker_row(
                sheet,
                columns,
                excel_row,
                candidate_id=candidate_id,
                offer_url=offer_url,
            )
            for header, kind in (("CV PDF", "cv"), ("Lettre PDF", "letter")):
                cell = sheet.cell(excel_row, columns[header])
                cell.value = header
                cell.hyperlink = str(paths[kind])
                cell.style = "Hyperlink"
            save_workbook_atomically(workbook, tracker_path)
        finally:
            workbook.close()


def _sync_completed_artifacts(
    record: StudioCampaignRecord,
    item: StudioCampaignItem,
    artifacts: dict[str, dict[str, object]],
) -> None:
    if item.excel_row is None:
        raise ValueError("completed campaign item has no tracker row")
    _sync_artifact_links(
        record.tracker_path,
        item.excel_row,
        artifacts,
        candidate_id=record.candidate_id,
        offer_url=item.offer.url,
    )
    item.tracker_artifacts_synced = True
    item.tracker_sync_error = None


def _find_tracker_row(
    tracker_path: Path,
    *,
    candidate_id: str,
    offer_url: str,
) -> int | None:
    target = canonical_url(offer_url)
    with tracker_lock(tracker_path):
        workbook = load_workbook(tracker_path, read_only=True, data_only=True)
        try:
            if "Postulations" not in workbook.sheetnames:
                return None
            sheet = workbook["Postulations"]
            columns = header_map(sheet)
            url_column = columns.get("Lien offre", 3)
            owner_column = columns.get(CANDIDATE_ID_COLUMN)
            for row in range(2, sheet.max_row + 1):
                tracked_url = str(sheet.cell(row, url_column).value or "").strip()
                if not tracked_url or canonical_url(tracked_url) != target:
                    continue
                owner = (
                    str(sheet.cell(row, owner_column).value or "").strip()
                    if owner_column is not None
                    else ""
                )
                if owner in {"", candidate_id}:
                    return row
            return None
        finally:
            workbook.close()


def _claim_and_validate_tracker_row(
    sheet,
    columns: dict[str, int],
    row: int,
    *,
    candidate_id: str,
    offer_url: str,
) -> None:
    owner_cell = sheet.cell(row, columns[CANDIDATE_ID_COLUMN])
    owner = str(owner_cell.value or "").strip()
    if owner and owner != candidate_id:
        raise ValueError(f"tracker row {row} belongs to another candidate")
    tracked_url = str(sheet.cell(row, columns["Lien offre"]).value or "").strip()
    if not tracked_url or canonical_url(tracked_url) != canonical_url(offer_url):
        raise ValueError(f"tracker row {row} no longer matches the campaign offer")
    owner_cell.value = candidate_id


def _atomic_write_json(path: Path, payload: dict[str, object]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        text=True,
    )
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_name, path)
    except BaseException:
        Path(temporary_name).unlink(missing_ok=True)
        raise
