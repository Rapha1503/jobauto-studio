from __future__ import annotations

import hashlib
import json
import os
import tempfile
import threading
from datetime import UTC, datetime
from pathlib import Path
from typing import Literal, Protocol

from openpyxl import load_workbook
from pydantic import BaseModel, ConfigDict, Field, model_validator

from jobauto.candidate_form_profile import CandidateFormProfile
from jobauto.candidate_profile import CandidateIdentity
from jobauto.candidate_snapshot import CandidateProfileRepository
from jobauto.excel_schema import CANDIDATE_ID_COLUMN, TRACKER_COLUMNS, ensure_tracker_schema
from jobauto.offer_catalog import canonical_url
from jobauto.run_store import RunRecord, utc_now
from jobauto.studio_campaign import StudioCampaignRecord, StudioCampaignService
from jobauto.submission_preferences import SubmissionMode, SubmissionPreferences
from jobauto.tracker_io import path_lock, save_workbook_atomically, tracker_lock

HandoffStatus = Literal[
    "ready_for_chrome",
    "claimed_for_chrome",
    "dry_run",
    "blocked",
    "sandbox_verified",
    "submitted",
]
ReceiptStatus = Literal["blocked", "sandbox_verified", "submitted"]


class RunReader(Protocol):
    def get(self, run_id: str) -> RunRecord: ...


class HandoffArtifact(BaseModel):
    model_config = ConfigDict(extra="forbid", frozen=True)

    kind: Literal["cv", "letter"]
    path: Path
    sha256: str = Field(pattern=r"^[0-9a-f]{64}$")
    page_count: int = Field(ge=1)


class SubmissionReceipt(BaseModel):
    model_config = ConfigDict(extra="forbid")

    status: ReceiptStatus
    portal: str = Field(min_length=1, max_length=200)
    submitted_at: str | None = None
    confirmation_url: str | None = None
    evidence_path: Path | None = None
    filled_fields: list[str] = Field(default_factory=list)
    uploaded_files: list[str] = Field(default_factory=list)
    blockers: list[str] = Field(default_factory=list)
    warnings: list[str] = Field(default_factory=list)

    @model_validator(mode="after")
    def terminal_receipt_has_timestamp(self) -> SubmissionReceipt:
        if self.status in {"submitted", "sandbox_verified"} and not self.submitted_at:
            self.submitted_at = datetime.now(UTC).isoformat()
        return self


class SubmissionHandoffRecord(BaseModel):
    model_config = ConfigDict(extra="forbid")

    handoff_id: str = Field(min_length=8, max_length=160)
    campaign_id: str
    run_id: str
    candidate_id: str
    candidate_identity: CandidateIdentity | None = None
    candidate_form_profile: CandidateFormProfile | None = None
    excel_row: int = Field(ge=2)
    company: str
    role: str
    offer_url: str
    status: HandoffStatus
    created_at: str
    updated_at: str
    artifacts: list[HandoffArtifact]
    preferences: SubmissionPreferences
    blockers: list[str] = Field(default_factory=list)
    receipt: SubmissionReceipt | None = None
    tracker_sync_error: str | None = None


class SubmissionHandoffStore:
    def __init__(self, root: Path) -> None:
        self.root = root.expanduser().resolve()
        self.root.mkdir(parents=True, exist_ok=True)

    def save(self, record: SubmissionHandoffRecord) -> SubmissionHandoffRecord:
        target = self.root / record.handoff_id
        target.mkdir(parents=True, exist_ok=True)
        record_path = target / "handoff.json"
        with path_lock(record_path):
            _atomic_write_json(record_path, record.model_dump(mode="json"))
        return record

    def get(self, handoff_id: str) -> SubmissionHandoffRecord:
        path = self.root / handoff_id / "handoff.json"
        if not path.is_file():
            raise FileNotFoundError(f"submission handoff not found: {handoff_id}")
        return SubmissionHandoffRecord.model_validate_json(path.read_text(encoding="utf-8"))

    def list_for_campaign(self, campaign_id: str) -> list[SubmissionHandoffRecord]:
        records: list[SubmissionHandoffRecord] = []
        for path in self.root.glob("*/handoff.json"):
            try:
                record = SubmissionHandoffRecord.model_validate_json(
                    path.read_text(encoding="utf-8")
                )
            except (OSError, ValueError):
                continue
            if record.campaign_id == campaign_id:
                records.append(record)
        return sorted(records, key=lambda record: (record.created_at, record.handoff_id))

    def find_for_run(
        self,
        campaign_id: str,
        run_id: str,
    ) -> SubmissionHandoffRecord | None:
        return next(
            (record for record in self.list_for_campaign(campaign_id) if record.run_id == run_id),
            None,
        )


class SubmissionHandoffService:
    def __init__(
        self,
        *,
        repository: CandidateProfileRepository,
        campaign_service: StudioCampaignService,
        run_reader: RunReader,
        store: SubmissionHandoffStore,
    ) -> None:
        self.repository = repository
        self.campaign_service = campaign_service
        self.run_reader = run_reader
        self.store = store
        self._claim_lock = threading.RLock()

    def prepare(
        self,
        campaign_id: str,
        run_id: str,
        *,
        preferences: SubmissionPreferences | None = None,
    ) -> SubmissionHandoffRecord:
        existing = self.store.find_for_run(campaign_id, run_id)
        campaign = self.campaign_service.get(campaign_id)
        item = next((entry for entry in campaign.items if entry.run_id == run_id), None)
        if item is None or item.excel_row is None:
            raise ValueError("run is not a selected item in this campaign")
        run = self.run_reader.get(run_id)
        snapshot = self.repository.load_snapshot(campaign.profile_path)
        form_profile = _load_form_profile(snapshot)
        if existing is not None:
            blockers = _run_blockers(run)
            artifacts = _verified_artifacts(run, blockers)
            return self.store.save(
                existing.model_copy(
                    update={
                        "candidate_form_profile": form_profile,
                        "artifacts": artifacts,
                        "blockers": list(dict.fromkeys([*existing.blockers, *blockers])),
                        "updated_at": utc_now(),
                    }
                )
            )
        blockers = _run_blockers(run)
        if snapshot.snapshot_hash != run.snapshot_hash:
            blockers.append("candidate_profile_changed_after_generation")
        artifacts = _verified_artifacts(run, blockers)
        selected_preferences = preferences or snapshot.submission_preferences
        status: HandoffStatus
        if blockers:
            status = "blocked"
        elif selected_preferences.mode is SubmissionMode.DRY_RUN:
            status = "dry_run"
        else:
            status = "ready_for_chrome"
        now = utc_now()
        record = SubmissionHandoffRecord(
            handoff_id=_handoff_id(campaign_id, run_id),
            campaign_id=campaign_id,
            run_id=run_id,
            candidate_id=campaign.candidate_id,
            candidate_identity=snapshot.profile.identity,
            candidate_form_profile=form_profile,
            excel_row=item.excel_row,
            company=item.offer.company,
            role=item.offer.role,
            offer_url=item.offer.url,
            status=status,
            created_at=now,
            updated_at=now,
            artifacts=artifacts,
            preferences=selected_preferences,
            blockers=blockers,
        )
        return self.store.save(record)

    def get(self, handoff_id: str) -> SubmissionHandoffRecord:
        return self.store.get(handoff_id)

    def set_mode(
        self,
        handoff_id: str,
        mode: SubmissionMode,
    ) -> SubmissionHandoffRecord:
        """Change an unclaimed queue packet between supervised and automatic apply."""
        record = self.store.get(handoff_id)
        if record.preferences.mode is mode:
            return record
        if record.status in {"claimed_for_chrome", "submitted"}:
            raise ValueError("submission mode cannot change after Chrome claimed the packet")
        preferences = record.preferences.model_copy(update={"mode": mode})
        if record.blockers:
            status: HandoffStatus = "blocked"
        elif mode is SubmissionMode.DRY_RUN:
            status = "dry_run"
        elif record.status == "sandbox_verified":
            status = "sandbox_verified"
        else:
            status = "ready_for_chrome"
        updated = record.model_copy(
            update={
                "preferences": preferences,
                "status": status,
                "updated_at": utc_now(),
            }
        )
        return self.store.save(updated)

    def claim_next(self, campaign_id: str) -> SubmissionHandoffRecord | None:
        """Atomically reserve one packet so two Codex tasks cannot submit it."""
        with self._claim_lock:
            record = next(
                (
                    item
                    for item in self.store.list_for_campaign(campaign_id)
                    if item.status in {"ready_for_chrome", "sandbox_verified"}
                ),
                None,
            )
            if record is None:
                return None
            claimed = record.model_copy(
                update={"status": "claimed_for_chrome", "updated_at": utc_now()}
            )
            return self.store.save(claimed)

    def release_claim(self, handoff_id: str) -> SubmissionHandoffRecord:
        with self._claim_lock:
            record = self.store.get(handoff_id)
            if record.status != "claimed_for_chrome":
                raise ValueError("only a claimed handoff can be released")
            restored_status: HandoffStatus = (
                "sandbox_verified"
                if record.receipt is not None and record.receipt.status == "sandbox_verified"
                else "ready_for_chrome"
            )
            released = record.model_copy(
                update={"status": restored_status, "updated_at": utc_now()}
            )
            return self.store.save(released)

    def record_receipt(
        self,
        handoff_id: str,
        receipt: SubmissionReceipt,
        *,
        allow_sandbox: bool = False,
    ) -> SubmissionHandoffRecord:
        record = self.store.get(handoff_id)
        if record.status == "blocked":
            raise ValueError("blocked handoff cannot receive a submission receipt")
        if record.receipt is not None and _same_receipt(record.receipt, receipt):
            if record.tracker_sync_error and receipt.status != "sandbox_verified":
                campaign = self.campaign_service.get(record.campaign_id)
                return self._sync_tracker(campaign, record)
            return record
        if record.status == "submitted":
            raise ValueError("submitted handoff cannot receive a conflicting receipt")
        if record.status == "dry_run" and receipt.status == "submitted":
            raise ValueError("dry-run handoff cannot receive an employer submission receipt")
        if receipt.status == "sandbox_verified" and not allow_sandbox:
            raise ValueError("sandbox receipt must come from the sandbox submission route")
        receipt = receipt.model_copy(
            update={
                "submitted_at": (
                    datetime.now(UTC).isoformat()
                    if receipt.status in {"submitted", "sandbox_verified"}
                    else None
                )
            }
        )
        _rehash_artifacts(record.artifacts)
        if (
            receipt.status == "submitted"
            and record.preferences.require_confirmation_evidence
            and not receipt.confirmation_url
            and receipt.evidence_path is None
        ):
            raise ValueError("submitted receipt requires confirmation evidence")
        if receipt.evidence_path is not None and not receipt.evidence_path.expanduser().is_file():
            raise ValueError("receipt evidence path does not exist")
        campaign = self.campaign_service.get(record.campaign_id)
        status: HandoffStatus = receipt.status
        updated = record.model_copy(
            update={
                "status": status,
                "updated_at": utc_now(),
                "receipt": receipt,
                "blockers": receipt.blockers,
            }
        )
        self.store.save(updated)
        if receipt.status != "sandbox_verified":
            updated = self._sync_tracker(campaign, updated)
        return updated

    def _sync_tracker(
        self,
        campaign: StudioCampaignRecord,
        record: SubmissionHandoffRecord,
    ) -> SubmissionHandoffRecord:
        try:
            _sync_receipt_to_tracker(campaign, record)
        except (OSError, KeyError, ValueError) as exc:
            updated = record.model_copy(
                update={"tracker_sync_error": f"{type(exc).__name__}: {exc}"}
            )
        else:
            updated = record.model_copy(update={"tracker_sync_error": None})
        return self.store.save(updated)


def _load_form_profile(snapshot) -> CandidateFormProfile:
    path = snapshot.profile.form_profile_path
    if path is not None:
        return CandidateFormProfile.load(path)
    return CandidateFormProfile.from_cv_source(snapshot.cv_source)


def _run_blockers(run: RunRecord) -> list[str]:
    blockers = list(run.blockers)
    if run.status != "completed":
        blockers.append(f"run_not_completed:{run.status}")
    if not run.review or not bool(run.review.get("approved")):
        blockers.append("final_review_not_approved")
    return list(dict.fromkeys(blockers))


def _verified_artifacts(run: RunRecord, blockers: list[str]) -> list[HandoffArtifact]:
    artifacts: list[HandoffArtifact] = []
    for kind in ("cv", "letter"):
        payload = run.artifacts.get(kind)
        if payload is None:
            blockers.append(f"{kind}_artifact_missing")
            continue
        path = Path(str(payload.get("pdf_path", ""))).expanduser().resolve()
        expected = str(payload.get("pdf_sha256", ""))
        if not path.is_file():
            blockers.append(f"{kind}_pdf_missing")
            continue
        actual = _file_sha256(path)
        if actual != expected:
            blockers.append(f"{kind}_hash_mismatch")
            continue
        artifacts.append(
            HandoffArtifact(
                kind=kind,
                path=path,
                sha256=actual,
                page_count=int(payload.get("page_count", 0)),
            )
        )
    return artifacts


def _rehash_artifacts(artifacts: list[HandoffArtifact]) -> None:
    for artifact in artifacts:
        if not artifact.path.is_file() or _file_sha256(artifact.path) != artifact.sha256:
            raise ValueError(f"{artifact.kind} artifact changed after Chrome handoff")


def _sync_receipt_to_tracker(
    campaign: StudioCampaignRecord,
    handoff: SubmissionHandoffRecord,
) -> None:
    receipt = handoff.receipt
    if receipt is None:
        return
    with tracker_lock(campaign.tracker_path):
        workbook = load_workbook(campaign.tracker_path)
        try:
            sheet = workbook["Postulations"]
            columns = ensure_tracker_schema(sheet)
            row = handoff.excel_row
            _claim_and_validate_tracker_row(sheet, columns, row, handoff)
            if receipt.status == "submitted":
                sheet.cell(row, columns[TRACKER_COLUMNS[11]]).value = "Oui"
                sheet.cell(row, columns[TRACKER_COLUMNS[12]]).value = receipt.submitted_at
                sheet.cell(row, columns[TRACKER_COLUMNS[13]]).value = "Postulé"
            sheet.cell(row, columns[TRACKER_COLUMNS[20]]).value = receipt.portal
            sheet.cell(row, columns[TRACKER_COLUMNS[21]]).value = receipt.status
            sheet.cell(row, columns[TRACKER_COLUMNS[22]]).value = receipt.submitted_at or utc_now()
            sheet.cell(row, columns[TRACKER_COLUMNS[23]]).value = "; ".join(
                [*receipt.blockers, *receipt.warnings]
            )
            sheet.cell(row, columns[TRACKER_COLUMNS[25]]).value = "chrome_extension"
            if receipt.evidence_path is not None:
                evidence = receipt.evidence_path.expanduser().resolve()
                cell = sheet.cell(row, columns[TRACKER_COLUMNS[24]])
                cell.value = "Confirmation"
                cell.hyperlink = str(evidence)
                cell.style = "Hyperlink"
            save_workbook_atomically(workbook, campaign.tracker_path)
        finally:
            workbook.close()


def _claim_and_validate_tracker_row(
    sheet,
    columns: dict[str, int],
    row: int,
    handoff: SubmissionHandoffRecord,
) -> None:
    owner_cell = sheet.cell(row, columns[CANDIDATE_ID_COLUMN])
    owner = str(owner_cell.value or "").strip()
    if owner and owner != handoff.candidate_id:
        raise ValueError(f"tracker row {row} belongs to another candidate")
    tracked_url = str(sheet.cell(row, columns["Lien offre"]).value or "").strip()
    if not tracked_url or canonical_url(tracked_url) != canonical_url(handoff.offer_url):
        raise ValueError(f"tracker row {row} no longer matches the handoff offer")
    owner_cell.value = handoff.candidate_id


def _file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as handle:
        for chunk in iter(lambda: handle.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def _handoff_id(campaign_id: str, run_id: str) -> str:
    digest = hashlib.sha256(f"{campaign_id}\0{run_id}".encode()).hexdigest()
    return f"handoff-{digest[:16]}"


def _same_receipt(left: SubmissionReceipt, right: SubmissionReceipt) -> bool:
    ignored = {"submitted_at"}
    left_payload = left.model_dump(mode="json", exclude=ignored)
    right_payload = right.model_dump(mode="json", exclude=ignored)
    return left_payload == right_payload


def _atomic_write_json(path: Path, payload: dict[str, object]) -> None:
    descriptor, temporary_name = tempfile.mkstemp(
        dir=path.parent,
        prefix=f".{path.name}.",
        suffix=".tmp",
        text=True,
    )
    try:
        with os.fdopen(descriptor, "w", encoding="utf-8", newline="\n") as handle:
            json.dump(payload, handle, ensure_ascii=False, indent=2)
            handle.write("\n")
            handle.flush()
            os.fsync(handle.fileno())
        os.replace(temporary_name, path)
    except BaseException:
        Path(temporary_name).unlink(missing_ok=True)
        raise
