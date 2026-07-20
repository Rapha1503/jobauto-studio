from __future__ import annotations

import json
import re
import unicodedata
from collections.abc import Iterable
from dataclasses import dataclass, field
from datetime import date, timedelta
from pathlib import Path
from urllib.parse import parse_qsl, urlencode, urlsplit, urlunsplit

from openpyxl import load_workbook
from pydantic import BaseModel, Field

from jobauto.excel_schema import (
    CANDIDATE_ID_COLUMN,
    TRACKER_COLUMNS,
    ensure_tracker_schema,
    header_map,
)
from jobauto.tracker_io import save_workbook_atomically, tracker_lock

MIN_FULL_DESCRIPTION_LENGTH = 200
_TRACKING_PARAMS = {"utm_source", "utm_medium", "utm_campaign", "utm_term", "utm_content"}

_FRENCH_MONTHS = {
    "janvier": 1,
    "fevrier": 2,
    "mars": 3,
    "avril": 4,
    "mai": 5,
    "juin": 6,
    "juillet": 7,
    "aout": 8,
    "septembre": 9,
    "octobre": 10,
    "novembre": 11,
    "decembre": 12,
}
_ENGLISH_MONTHS = {
    "january": 1,
    "february": 2,
    "march": 3,
    "april": 4,
    "may": 5,
    "june": 6,
    "july": 7,
    "august": 8,
    "september": 9,
    "october": 10,
    "november": 11,
    "december": 12,
}


class OfferCandidate(BaseModel):
    company: str = Field(min_length=1)
    role: str = Field(min_length=1)
    url: str = Field(min_length=8)
    posted_at: str | None = None
    posted_text: str | None = None
    location: str | None = None
    language: str | None = None
    description: str = ""
    source: str = "candidate"
    experience_required: str | None = None
    contract_type: str | None = None
    salary_estimate: str | None = None
    semantic_fit_score: int | None = Field(default=None, ge=0, le=100)
    semantic_fit_rationale: str | None = Field(default=None, max_length=500)


@dataclass(frozen=True)
class SelectedOffer:
    company: str
    role: str
    url: str
    posted_at: str | None
    location: str | None
    description: str
    fit_score: int
    reasons: list[str] = field(default_factory=list)
    warnings: list[str] = field(default_factory=list)
    source: str = "candidate"
    experience_required: str | None = None
    contract_type: str | None = None
    salary_estimate: str | None = None
    semantic_fit_score: int | None = None
    semantic_fit_rationale: str | None = None

    def as_candidate(self) -> OfferCandidate:
        return OfferCandidate(
            company=self.company,
            role=self.role,
            url=self.url,
            posted_at=self.posted_at,
            location=self.location,
            description=self.description,
            source=self.source,
            experience_required=self.experience_required,
            contract_type=self.contract_type,
            salary_estimate=self.salary_estimate,
            semantic_fit_score=self.semantic_fit_score,
            semantic_fit_rationale=self.semantic_fit_rationale,
        )


def canonical_url(url: str) -> str:
    parts = urlsplit(url.strip())
    query = [
        (key, value)
        for key, value in parse_qsl(parts.query)
        if key.casefold() not in _TRACKING_PARAMS
    ]
    return urlunsplit(
        (parts.scheme.lower(), parts.netloc.lower(), parts.path.rstrip("/"), urlencode(query), "")
    )


def load_candidates(path: Path) -> list[OfferCandidate]:
    raw = json.loads(path.read_text(encoding="utf-8-sig"))
    if isinstance(raw, dict):
        raw = raw.get("offers") or raw.get("candidates") or []
    if not isinstance(raw, list):
        return []
    candidates: list[OfferCandidate] = []
    for item in raw:
        if not isinstance(item, dict):
            continue
        normalized = {
            "company": item.get("company") or item.get("entreprise") or item.get("employer"),
            "role": item.get("role") or item.get("title") or item.get("poste"),
            "url": _clean_candidate_url(item.get("url") or item.get("link") or item.get("lien")),
            "posted_at": item.get("posted_at") or item.get("date"),
            "posted_text": item.get("posted_text") or item.get("posted"),
            "location": item.get("location") or item.get("lieu"),
            "language": item.get("language") or item.get("langue"),
            "description": item.get("description")
            or item.get("summary")
            or item.get("notes")
            or "",
            "source": item.get("source") or "candidate_file",
            "experience_required": item.get("experience_required")
            or item.get("xp")
            or item.get("xp_demandee")
            or item.get("xp demandée"),
            "contract_type": item.get("contract_type")
            or item.get("contrat")
            or item.get("type_contrat")
            or item.get("type contrat"),
            "semantic_fit_score": item.get("semantic_fit_score"),
            "semantic_fit_rationale": item.get("semantic_fit_rationale"),
            "salary_estimate": item.get("salary_estimate")
            or item.get("salaire")
            or item.get("salaire_estime")
            or item.get("salaire estimé"),
        }
        if normalized["company"] and normalized["role"] and normalized["url"]:
            candidates.append(OfferCandidate(**normalized))
    return candidates


def append_selected_offers(
    workbook_path: Path,
    offers: Iterable[OfferCandidate | SelectedOffer],
    *,
    sheet_name: str = "Postulations",
    candidate_id: str | None = None,
) -> list[int]:
    with tracker_lock(workbook_path):
        workbook = load_workbook(workbook_path)
        try:
            sheet = (
                workbook[sheet_name]
                if sheet_name in workbook.sheetnames
                else workbook.create_sheet(sheet_name)
            )
            columns = ensure_tracker_schema(sheet)
            rows: list[int] = []
            for offer in offers:
                candidate = offer.as_candidate() if isinstance(offer, SelectedOffer) else offer
                row = _next_empty_row(sheet)
                values = {
                    TRACKER_COLUMNS[0]: candidate.company,
                    TRACKER_COLUMNS[1]: candidate.role,
                    TRACKER_COLUMNS[2]: candidate.url,
                    TRACKER_COLUMNS[5]: (
                        candidate.description.strip()
                        if len(candidate.description.strip()) >= MIN_FULL_DESCRIPTION_LENGTH
                        else None
                    ),
                    TRACKER_COLUMNS[6]: candidate.posted_at,
                    TRACKER_COLUMNS[7]: candidate.location,
                    TRACKER_COLUMNS[8]: candidate.experience_required,
                    TRACKER_COLUMNS[9]: candidate.contract_type,
                    TRACKER_COLUMNS[10]: candidate.salary_estimate,
                    TRACKER_COLUMNS[17]: _priority_stars(offer),
                    TRACKER_COLUMNS[18]: candidate.source,
                    TRACKER_COLUMNS[19]: date.today().isoformat(),
                    CANDIDATE_ID_COLUMN: candidate_id,
                }
                for header, value in values.items():
                    if value is not None and str(value).strip():
                        sheet.cell(row, columns[header]).value = str(value).strip()
                rows.append(row)
            save_workbook_atomically(workbook, workbook_path)
            return rows
        finally:
            workbook.close()


def existing_urls_from_workbook(
    workbook_path: Path,
    *,
    sheet_name: str = "Postulations",
    candidate_id: str | None = None,
) -> set[str]:
    with tracker_lock(workbook_path):
        workbook = load_workbook(workbook_path, read_only=True, data_only=True)
        try:
            if sheet_name not in workbook.sheetnames:
                return set()
            sheet = workbook[sheet_name]
            columns = header_map(sheet)
            url_column = columns.get(TRACKER_COLUMNS[2], 3)
            owner_column = columns.get(CANDIDATE_ID_COLUMN)
            urls: set[str] = set()
            for row in range(2, sheet.max_row + 1):
                owner = (
                    str(sheet.cell(row, owner_column).value or "").strip()
                    if owner_column is not None
                    else ""
                )
                if candidate_id is not None and owner and owner != candidate_id:
                    continue
                value = sheet.cell(row, url_column).value
                if value:
                    urls.add(canonical_url(str(value)))
            return urls
        finally:
            workbook.close()


def parse_posted_date(candidate: OfferCandidate, *, today: date) -> date | None:
    return parse_posted_date_text(
        " ".join(
            value
            for value in (
                candidate.posted_at,
                candidate.posted_text,
                candidate.description[:2000],
            )
            if value
        ),
        today=today,
    )


def parse_posted_date_text(text: str, *, today: date | None = None) -> date | None:
    today = today or date.today()
    normalized = _text_key(text)
    iso = re.search(r"(?<!\d)(20\d{2})-(\d{2})-(\d{2})(?!\d)", normalized)
    if iso:
        return date(int(iso.group(1)), int(iso.group(2)), int(iso.group(3)))
    relative = re.search(r"\b(\d{1,2})\+?\s*(?:days?|jours?)\s*(?:ago|il y a)?\b", normalized)
    if relative:
        return today - timedelta(days=int(relative.group(1)))
    for pattern, day_group, month_group, year_group in (
        (r"\b([a-z]+)\s+(\d{1,2}),?\s+(20\d{2})\b", 2, 1, 3),
        (r"\b(\d{1,2})\s+([a-z]+)\s+(20\d{2})\b", 1, 2, 3),
    ):
        for match in re.finditer(pattern, normalized.replace(".", "")):
            month_name = match.group(month_group)
            month = _FRENCH_MONTHS.get(month_name) or _ENGLISH_MONTHS.get(month_name)
            if month:
                return date(
                    int(match.group(year_group)),
                    month,
                    int(match.group(day_group)),
                )
    return None


def _clean_candidate_url(value: object) -> str | None:
    text = str(value or "").strip()
    if not text:
        return None
    match = re.search(r"https?://[^\s)]+", text)
    return match.group(0).rstrip(".,;") if match else text


def _priority_stars(offer: OfferCandidate | SelectedOffer) -> str | None:
    if not isinstance(offer, SelectedOffer):
        return None
    return "★★★" if offer.fit_score >= 85 else "★★" if offer.fit_score >= 75 else "★"


def _next_empty_row(sheet) -> int:
    for row in range(2, sheet.max_row + 2):
        if not any(sheet.cell(row, column).value for column in range(1, 4)):
            return row
    return sheet.max_row + 1


def _text_key(value: str) -> str:
    decomposed = unicodedata.normalize("NFKD", value.casefold())
    accentless = "".join(char for char in decomposed if not unicodedata.combining(char))
    return " ".join(accentless.split())
