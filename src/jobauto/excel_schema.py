from __future__ import annotations

from copy import copy

from openpyxl.worksheet.worksheet import Worksheet

CANDIDATE_ID_COLUMN = "Candidate ID"

TRACKER_COLUMNS = [
    "Entreprise",
    "Poste",
    "Lien offre",
    "CV PDF",
    "Lettre PDF",
    "Description",
    "Date publication",
    "Lieu",
    "XP demandée",
    "Type contrat",
    "Salaire estimé",
    "Postulé ?",
    "Date postulation",
    "Statut",
    "Canal",
    "Contact RH / Nom",
    "Notes / Relances",
    "Priorité",
    "Source recherche",
    "Date ajout",
    "Portail candidature",
    "Statut candidature",
    "Dernier essai candidature",
    "Notes candidature",
    "Screenshot pre-submit",
    "Outil automation",
    CANDIDATE_ID_COLUMN,
]

RETIRED_COLUMNS = [
    "Dossier candidature",
    "Candidature ouverte",
    "Réalisme / Fit",
]


def header_map(sheet: Worksheet) -> dict[str, int]:
    return {
        str(sheet.cell(1, column).value or "").strip(): column
        for column in range(1, sheet.max_column + 1)
        if str(sheet.cell(1, column).value or "").strip()
    }


def ensure_tracker_schema(sheet: Worksheet) -> dict[str, int]:
    for header in RETIRED_COLUMNS:
        current = header_map(sheet)
        source_index = current.get(header)
        if source_index is not None:
            sheet.delete_cols(source_index)
    for target_index, header in enumerate(TRACKER_COLUMNS, start=1):
        current = header_map(sheet)
        source_index = current.get(header)
        if source_index is None:
            sheet.insert_cols(target_index)
            sheet.cell(1, target_index).value = header
            _set_default_width(sheet, target_index, header)
        elif source_index != target_index:
            _move_column(sheet, source_index, target_index)
    for target in ("CV PDF", "Lettre PDF", "Screenshot pre-submit"):
        _sync_header_style(sheet, target, "Lien offre")
    _sync_header_style(sheet, "Outil automation", "Screenshot pre-submit")
    current = header_map(sheet)
    for header, column in current.items():
        if header in TRACKER_COLUMNS:
            _set_default_width(sheet, column, header)
    _normalize_sheet_view(sheet)
    candidate_column = header_map(sheet).get(CANDIDATE_ID_COLUMN)
    if candidate_column is not None:
        sheet.column_dimensions[sheet.cell(1, candidate_column).column_letter].hidden = True
    return header_map(sheet)


def _move_column(sheet: Worksheet, source: int, target: int) -> None:
    if source == target:
        return
    if source > target:
        sheet.insert_cols(target)
        _copy_column(sheet, source + 1, target)
        sheet.delete_cols(source + 1)
    else:
        sheet.insert_cols(target + 1)
        _copy_column(sheet, source, target + 1)
        sheet.delete_cols(source)
    _set_default_width(sheet, target, str(sheet.cell(1, target).value or ""))


def _copy_column(sheet: Worksheet, source: int, target: int) -> None:
    source_letter = sheet.cell(1, source).column_letter
    target_letter = sheet.cell(1, target).column_letter
    if source_letter in sheet.column_dimensions:
        sheet.column_dimensions[target_letter].width = sheet.column_dimensions[source_letter].width
    for row in range(1, sheet.max_row + 1):
        source_cell = sheet.cell(row, source)
        target_cell = sheet.cell(row, target)
        target_cell.value = source_cell.value
        if source_cell.has_style:
            target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.hyperlink = copy(source_cell.hyperlink)
        target_cell.comment = copy(source_cell.comment)


def _set_default_width(sheet: Worksheet, column: int, header: str) -> None:
    widths = {
        "Lien offre": 36,
        "CV PDF": 18,
        "Lettre PDF": 18,
        "Description": 55,
        "Salaire estimé": 18,
        "Notes / Relances": 36,
        "Notes candidature": 36,
        "Screenshot pre-submit": 22,
        "Outil automation": 22,
        CANDIDATE_ID_COLUMN: 24,
    }
    sheet.column_dimensions[sheet.cell(1, column).column_letter].width = widths.get(header, 16)


def _sync_header_style(sheet: Worksheet, target_header: str, reference_header: str) -> None:
    columns = header_map(sheet)
    target = columns.get(target_header)
    reference = columns.get(reference_header)
    if not target or not reference:
        return
    source_cell = sheet.cell(1, reference)
    target_cell = sheet.cell(1, target)
    if source_cell.has_style:
        target_cell._style = copy(source_cell._style)
    target_cell.font = copy(source_cell.font)
    target_cell.fill = copy(source_cell.fill)
    target_cell.border = copy(source_cell.border)
    target_cell.alignment = copy(source_cell.alignment)
    target_cell.number_format = source_cell.number_format
    target_cell.protection = copy(source_cell.protection)


def _normalize_sheet_view(sheet: Worksheet) -> None:
    sheet.freeze_panes = "A2"
