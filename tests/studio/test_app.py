from __future__ import annotations

import base64
import hashlib
import json
import shutil
from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path
from types import SimpleNamespace

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from jobauto.candidate_draft import CandidateDraftUpdate
from jobauto.candidate_profile import CvBackend
from jobauto.candidate_snapshot import CandidateProfileRepository
from jobauto.document_patch import (
    CvAdaptationPatch,
    CvFieldChange,
    apply_cv_patch,
    editable_cv_source_index,
)
from jobauto.excel_schema import CANDIDATE_ID_COLUMN, TRACKER_COLUMNS
from jobauto.profile_extraction import CandidateProfileExtraction
from jobauto.run_store import RunRecord
from jobauto.studio.app import (
    _ensure_default_tracker,
    _installed_codex_plugins,
    create_studio_app,
)
from jobauto.studio.tex_imports import TexImportStore


def _client(tmp_path: Path) -> TestClient:
    project_root = Path(__file__).resolve().parents[2]
    return TestClient(create_studio_app(project_root=project_root, state_root=tmp_path))


def test_default_tracker_initialization_is_safe_under_concurrent_calls(tmp_path: Path) -> None:
    tracker = tmp_path / "state" / "applications.xlsx"

    with ThreadPoolExecutor(max_workers=4) as executor:
        list(executor.map(lambda _index: _ensure_default_tracker(tracker), range(8)))

    workbook = load_workbook(tracker, read_only=False, data_only=True)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        assert CANDIDATE_ID_COLUMN in headers
        column_letter = sheet.cell(1, headers[CANDIDATE_ID_COLUMN]).column_letter
        assert sheet.column_dimensions[column_letter].hidden is True
    finally:
        workbook.close()


def test_studio_home_shows_safe_example_profile(tmp_path: Path) -> None:
    response = _client(tmp_path).get("/")

    assert response.status_code == 200
    assert "JobAuto Studio" in response.text
    assert "gpt-5.6-sol" in response.text
    assert "Alex Morgan" in response.text
    assert "PrivateCandidate" not in response.text
    assert "Create my JobAuto profile" in response.text
    hero_start = response.text.index('<section class="hero home-hero">')
    hero_end = response.text.index("</section>", hero_start)
    assert 'class="hero-action" href="/setup"' in response.text[hero_start:hero_end]
    assert 'class="hero-secondary" href="/demo"' in response.text[hero_start:hero_end]
    assert ">Explore the checked demo</a>" in response.text[hero_start:hero_end]
    assert "Choose a candidate profile" not in response.text


def test_studio_home_source_default_loads_public_examples(tmp_path: Path) -> None:
    response = TestClient(create_studio_app(state_root=tmp_path)).get("/")

    assert response.status_code == 200
    assert "Alex Morgan" in response.text
    assert ">Explore the checked demo</a>" in response.text


def test_checked_demo_replays_non_technical_campaign_and_serves_artifacts(
    tmp_path: Path,
) -> None:
    client = _client(tmp_path)

    response = client.get("/demo")

    assert response.status_code == 200
    assert "Maya Laurent's application campaign" in response.text
    assert "Verified non-technical replay" in response.text
    assert "Five reviewed document packs" in response.text
    assert "Why this application matched" in response.text
    assert "End-to-end event logistics and on-site delivery" in response.text
    assert "4 offers kept outside this campaign" in response.text
    assert "below the five-offer campaign cutoff" in response.text
    assert "Event Studio Meridian" in response.text
    assert "UMDH - Une Marque & Des Hommes" not in response.text
    assert response.text.count("sandbox verified") == 5
    assert "Legacy Private Candidate" not in response.text
    assert "Legacy Private Employer" not in response.text

    source_cv = client.get("/demo/files/source-cv.pdf")
    tailored_cv = client.get("/demo/files/artifacts/01-umdh/cv.pdf")
    assert source_cv.status_code == 200
    assert tailored_cv.status_code == 200
    assert source_cv.headers["content-type"] == "application/pdf"
    assert tailored_cv.headers["content-type"] == "application/pdf"
    assert source_cv.content.startswith(b"%PDF")
    assert tailored_cv.content.startswith(b"%PDF")


def test_checked_demo_rejects_artifact_path_traversal(tmp_path: Path) -> None:
    response = _client(tmp_path).get("/demo/files/%2e%2e%2fREADME.md")

    assert response.status_code == 404


def test_studio_home_can_select_archive_and_restore_user_profiles(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[2]
    candidate_root = tmp_path / "candidate-profiles"
    alex = candidate_root / "user-alex"
    jamie = candidate_root / "user-jamie"
    shutil.copytree(project_root / "config" / "profiles" / "example", alex)
    shutil.copytree(project_root / "config" / "profiles" / "example-b", jamie)
    (alex / "profile.yaml").write_text(
        (alex / "profile.yaml")
        .read_text(encoding="utf-8")
        .replace("candidate_id: alex-morgan", "candidate_id: user-alex"),
        encoding="utf-8",
    )
    (jamie / "profile.yaml").write_text(
        (jamie / "profile.yaml")
        .read_text(encoding="utf-8")
        .replace("candidate_id: jamie-chen", "candidate_id: user-jamie"),
        encoding="utf-8",
    )
    client = _client(tmp_path)

    home = client.get("/")
    assert home.status_code == 200
    assert "Choose a profile" in home.text
    assert 'data-select-profile="user-alex"' in home.text
    assert 'data-select-profile="user-jamie"' in home.text

    selected = client.post("/profiles/user-alex/select")
    assert selected.status_code == 200
    assert selected.json()["page_url"] == "/profiles/user-alex"
    state = json.loads((tmp_path / "profile-workspace.json").read_text(encoding="utf-8"))
    assert state["active_candidate_id"] == "user-alex"

    archived = client.post("/profiles/user-alex/archive")
    assert archived.status_code == 200
    home = client.get("/")
    assert "Archived profiles" in home.text
    assert 'data-restore-profile="user-alex"' in home.text
    assert 'data-select-profile="user-alex"' not in home.text

    restored = client.post("/profiles/user-alex/restore")
    assert restored.status_code == 200
    assert 'data-select-profile="user-alex"' in client.get("/").text
    profile = client.get("/profiles/user-alex")
    assert 'id="start-over-profile"' in profile.text
    assert 'id="archive-profile"' in profile.text


def test_studio_home_shows_only_latest_version_of_the_same_draft(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[2]
    candidate_root = tmp_path / "candidate-profiles"
    for version in (1, 2):
        target = candidate_root / f"alex-v{version}"
        shutil.copytree(project_root / "config" / "profiles" / "example", target)
        profile_path = target / "profile.yaml"
        profile_path.write_text(
            profile_path.read_text(encoding="utf-8").replace(
                "candidate_id: alex-morgan", f"candidate_id: alex-morgan-v{version}"
            ),
            encoding="utf-8",
        )
        (target / "studio_source.json").write_text(
            json.dumps({"draft_id": "same-draft"}),
            encoding="utf-8",
        )

    home = _client(tmp_path).get("/")

    assert home.status_code == 200
    assert 'data-select-profile="alex-morgan-v1"' not in home.text
    assert 'data-select-profile="alex-morgan-v2"' in home.text


def test_run_page_retries_tailored_preview_after_artifacts_are_ready() -> None:
    template = (
        Path(__file__).resolve().parents[2]
        / "src"
        / "jobauto"
        / "studio"
        / "templates"
        / "run.html"
    ).read_text(encoding="utf-8")

    assert "tailoredPreview.dataset.ready !== 'true'" in template
    assert "previews/cv.png?ready=" in template


def test_setup_reports_local_runtime_requirements(tmp_path: Path, monkeypatch) -> None:
    monkeypatch.setattr(
        "jobauto.studio.app.shutil.which",
        lambda command: "C:/tools/codex.exe" if command == "codex" else None,
    )
    monkeypatch.setattr(
        "jobauto.studio.app._installed_codex_plugins",
        lambda: {"chrome@openai-bundled"},
    )

    response = _client(tmp_path).get("/setup")

    assert response.status_code == 200
    assert 'data-requirement="Codex CLI"' in response.text
    assert 'data-requirement="LaTeX"' in response.text
    assert 'data-requirement="Codex Chrome control"' in response.text
    assert 'data-requirement="JobAuto plugin"' in response.text
    assert 'data-ready="true"' in response.text
    assert 'data-ready="false"' in response.text
    assert "offer discovery and document agents" in response.text
    assert "source preview and final PDFs" in response.text
    assert "codex plugin marketplace add ." in response.text
    assert "codex plugin add jobauto@jobauto-studio" in response.text


def test_installed_codex_plugins_reads_only_enabled_installations(monkeypatch) -> None:
    monkeypatch.setattr("jobauto.studio.app.shutil.which", lambda _command: "codex.exe")
    monkeypatch.setattr(
        "jobauto.studio.app.subprocess.run",
        lambda *_args, **_kwargs: SimpleNamespace(
            returncode=0,
            stdout=json.dumps(
                {
                    "installed": [
                        {
                            "pluginId": "chrome@openai-bundled",
                            "installed": True,
                            "enabled": True,
                        },
                        {
                            "pluginId": "jobauto@jobauto-studio",
                            "installed": True,
                            "enabled": False,
                        },
                    ]
                }
            ),
        ),
    )

    assert _installed_codex_plugins() == {"chrome@openai-bundled"}


def test_profile_extraction_uses_the_selected_codex_model(tmp_path: Path, monkeypatch) -> None:
    seen: dict[str, object] = {}

    class FakeCodexClient:
        def complete_json(self, prompt, response_model, phase):
            blocks = json.loads(prompt.split("CONFIRMED_CV_BLOCKS_JSON\n", 1)[1])
            block_id = blocks[0]["block_id"]
            return CandidateProfileExtraction.model_validate(
                {
                    "identity": {
                        "first_name": "Alex",
                        "last_name": "Morgan",
                        "email": "alex.morgan@example.test",
                        "source_block_ids": [block_id],
                    }
                }
            )

    def fake_default(*, cwd, model=None, event_callback=None):
        seen.update({"cwd": cwd, "model": model, "event_callback": event_callback})
        return FakeCodexClient()

    monkeypatch.setattr("jobauto.studio.app.CodexClient.default", fake_default)
    project_root = Path(__file__).resolve().parents[2]
    client = TestClient(
        create_studio_app(
            project_root=project_root,
            state_root=tmp_path,
            codex_model="gpt-5.6-sol",
        )
    )
    source = rb"""\documentclass{article}
\begin{document}
Alex Morgan\\
alex.morgan@example.test
\end{document}
"""
    imported = client.post(
        "/api/tex-imports",
        content=source,
        headers={"X-Filename": "alex.tex", "Content-Type": "application/x-tex"},
    ).json()

    response = client.post(f"/api/tex-imports/{imported['import_id']}/profile-draft")

    assert response.status_code == 202
    assert seen["model"] == "gpt-5.6-sol"


def test_studio_templates_do_not_contain_mojibake_markers() -> None:
    studio_root = Path(__file__).resolve().parents[2] / "src" / "jobauto" / "studio"

    for path in [*studio_root.rglob("*.html"), *studio_root.rglob("*.css")]:
        content = path.read_text(encoding="utf-8")
        assert not any(marker in content for marker in ("Â", "Ã", "â€")), path


def test_studio_uses_an_internal_tracker_when_user_does_not_choose_one(tmp_path: Path) -> None:
    client = _client(tmp_path)
    assert client.get("/").status_code == 200
    assert (tmp_path / "applications.xlsx").is_file()


def test_setup_imports_real_tex_and_displays_detected_mapping(tmp_path: Path) -> None:
    client = _client(tmp_path)
    source = rb"""\documentclass{article}
\newcommand{\cvsection}[1]{\section*{#1}}
\begin{document}
Alex Morgan
\cvsection{Profile}
Data engineer building reliable pipelines.
\cvsection{Skills}
Python, SQL
\end{document}
"""

    setup = client.get("/setup")
    assert setup.status_code == 200
    assert "Bring your CV, or build it here" in setup.text
    assert "Build my CV in JobAuto" in setup.text

    response = client.post(
        "/api/tex-imports",
        content=source,
        headers={"X-Filename": "alex.tex", "Content-Type": "application/x-tex"},
    )

    assert response.status_code == 201
    payload = response.json()
    assert payload["compilation_status"] == "compiled"
    page = client.get(payload["page_url"])
    assert page.status_code == 200
    assert "CV imported" in page.text
    assert "Profile" in page.text
    assert "Skills" in page.text
    assert "Balanced" in page.text
    assert "Advanced: customize a specific section" in page.text
    assert "Add block" in page.text
    preview = client.get(f"{payload['page_url']}/preview.pdf")
    assert preview.status_code == 200
    assert preview.content.startswith(b"%PDF")
    preview_image = client.get(f"{payload['page_url']}/preview.png")
    assert preview_image.status_code == 200
    assert preview_image.headers["content-type"] == "image/png"
    assert preview_image.content.startswith(b"\x89PNG\r\n\x1a\n")

    mapping = TexImportStore(tmp_path / "tex-imports").mapping(payload["import_id"])
    summary = next(block for block in mapping.blocks if block.kind.value == "summary")
    saved = client.post(
        f"/api/tex-imports/{payload['import_id']}/mapping",
        json={
            "blocks": [
                {
                    "block_id": summary.block_id,
                    "label": summary.label,
                    "kind": summary.kind.value,
                    "start_line": summary.start_line,
                    "end_line": summary.end_line,
                    "fidelity": "very_faithful",
                    "required": True,
                    "target_lines": 3,
                }
            ]
        },
    )
    assert saved.status_code == 200
    updated = TexImportStore(tmp_path / "tex-imports").mapping(payload["import_id"])
    assert updated.blocks[0].policy.target_lines == 3


def test_manual_setup_exports_the_same_generated_profile_pipeline(tmp_path: Path) -> None:
    client = _client(tmp_path)
    created = client.post("/api/candidate-drafts/manual", json={"locale": "en-GB"})

    assert created.status_code == 201
    draft_id = created.json()["draft_id"]
    page = client.get(created.json()["page_url"])
    assert page.status_code == 200
    assert "Blank CV workspace" in page.text
    assert "Publications, certifications or anything else" in page.text

    draft = client.get(f"/api/candidate-drafts/{draft_id}").json()
    update = {
        name: draft[name]
        for name in CandidateDraftUpdate.model_fields
        if name != "expected_version" and draft.get(name) is not None
    }
    update.update(
        {
            "expected_version": draft["version"],
            "identity": {
                "first_name": "Jordan",
                "last_name": "Lee",
                "email": "jordan.lee@example.test",
                "location": "London",
                "headline": "Regulatory Affairs Specialist | Medical Devices",
                "source_block_ids": [],
            },
            "summary": (
                "Regulatory affairs specialist experienced in technical documentation, "
                "quality systems and cross-functional submissions for medical devices."
            ),
            "experiences": [
                {
                    "experience_id": "northbridge",
                    "organization": "Northbridge Health",
                    "role": "Regulatory Affairs Associate",
                    "dates": "2023 - 2026",
                    "sector": "Medical devices",
                    "tools": ["ISO 13485", "EU MDR"],
                    "facts": [
                        "Prepared technical documentation and coordinated submission evidence.",
                        "Worked with quality, clinical and engineering stakeholders.",
                    ],
                    "metrics": [],
                    "protected_fields": ["organization", "role", "dates"],
                    "allowed_angles": ["regulatory", "quality", "coordination"],
                    "source_block_ids": [],
                }
            ],
            "skills": [
                {
                    "name": "EU MDR",
                    "category": "Regulatory & Quality",
                    "usage": "default",
                    "evidence": "verified",
                    "verification_warning": False,
                    "source_block_ids": [],
                },
                {
                    "name": "ISO 13485",
                    "category": "Regulatory & Quality",
                    "usage": "default",
                    "evidence": "verified",
                    "verification_warning": False,
                    "source_block_ids": [],
                },
            ],
            "education": [
                {
                    "institution": "Kingston University",
                    "program": "MSc Biomedical Engineering",
                    "dates": "2021 - 2023",
                    "details": ["Medical-device design, risk and clinical evaluation."],
                    "source_block_ids": [],
                }
            ],
            "additional_sections": [
                {
                    "label": "Certifications",
                    "content": "ISO 13485 internal auditor training | Technical writing",
                    "fidelity": "adaptable",
                    "source_block_ids": [],
                }
            ],
            "languages": ["English native", "French B2"],
        }
    )
    saved = client.put(f"/api/candidate-drafts/{draft_id}", json=update)
    assert saved.status_code == 200, saved.text
    validated = client.post(f"/api/candidate-drafts/{draft_id}/validate")
    assert validated.status_code == 200, validated.text
    exported = client.post(f"/api/candidate-drafts/{draft_id}/export")
    assert exported.status_code == 201, exported.text

    snapshot = CandidateProfileRepository(tmp_path / "candidate-profiles").load_snapshot(
        Path(exported.json()["profile_path"])
    )
    assert snapshot.profile.cv_backend is CvBackend.GENERATED_TEMPLATE
    assert snapshot.cv_mapping is None
    assert snapshot.cv_source.additional_sections[0].label == "Certifications"
    assert "additional.0.content" in editable_cv_source_index(snapshot)
    adapted = apply_cv_patch(
        snapshot,
        CvAdaptationPatch(
            changes=[
                CvFieldChange(
                    source_id="additional.0.content",
                    value="ISO 13485 internal auditor training and technical writing",
                    fact_ids=["additional.1"],
                )
            ]
        ),
    )
    assert "technical writing" in adapted.document.additional_sections[0].content
    assert "EU MDR" in snapshot.cv_source.skills["Regulatory & Quality"]
    if shutil.which("pdflatex") is not None:
        preview = client.get(exported.json()["preview_url"])
        assert preview.status_code == 200, preview.text
        assert preview.content.startswith(b"%PDF")


def test_setup_rejects_non_tex_source(tmp_path: Path) -> None:
    response = _client(tmp_path).post(
        "/api/tex-imports",
        content=b"not latex",
        headers={"X-Filename": "resume.txt"},
    )

    assert response.status_code == 422
    assert "end with .tex" in response.json()["detail"]


def test_setup_rejects_tex_larger_than_upload_limit(tmp_path: Path) -> None:
    response = _client(tmp_path).post(
        "/api/tex-imports",
        content=b"x" * 2_000_001,
        headers={"X-Filename": "resume.tex"},
    )

    assert response.status_code == 413


class FakeProfileExtractor:
    def extract(self, _source, _mapping) -> CandidateProfileExtraction:
        return CandidateProfileExtraction.model_validate(
            {
                "identity": {
                    "first_name": "Camille",
                    "last_name": "Martin",
                    "email": "camille.martin@example.test",
                    "phone": "+33 1 00 00 00 01",
                    "location": "Lyon",
                    "headline": "Ingénieure systèmes",
                    "source_block_ids": ["identity"],
                },
                "summary": "Profil données et systèmes énergétiques.",
                "summary_source_block_ids": ["summary"],
                "experiences": [
                    {
                        "experience_id": "gridlab",
                        "organization": "GridLab",
                        "role": "Ingénieure données",
                        "dates": "2024--2026",
                        "tools": ["Python", "SQL"],
                        "facts": ["Développement de pipelines de données."],
                        "source_block_ids": ["experience"],
                    }
                ],
                "projects": [
                    {
                        "project_id": "energy_forecasting",
                        "title": "Prévision de demande énergétique",
                        "stack": ["Python", "scikit-learn"],
                        "description": ["Modélisation de séries temporelles."],
                        "source_block_ids": ["projects"],
                    }
                ],
                "skills": [
                    {
                        "name": "Python",
                        "category": "Data",
                        "source_block_ids": ["skills"],
                    }
                ],
                "education": [
                    {
                        "institution": "École du numérique",
                        "program": "Diplôme d'ingénieur",
                        "dates": "2021--2026",
                        "details": ["Machine learning et systèmes distribués."],
                        "source_block_ids": ["education"],
                    }
                ],
                "languages": ["Français natif", "Anglais C1"],
                "interests": ["Énergie"],
            }
        )


def test_confirmed_tex_can_become_an_observable_candidate_draft(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[2]
    client = TestClient(
        create_studio_app(
            project_root=project_root,
            state_root=tmp_path,
            profile_extractor=FakeProfileExtractor(),
        )
    )
    fixture = project_root / "tests" / "fixtures" / "cv" / "synthetic_cv_fr.tex"
    imported = client.post(
        "/api/tex-imports",
        content=fixture.read_bytes(),
        headers={"X-Filename": fixture.name},
    ).json()

    started = client.post(f"/api/tex-imports/{imported['import_id']}/profile-draft")
    status = client.get(started.json()["status_url"])

    assert started.status_code == 202
    assert status.json()["status"] == "completed"
    page = client.get(status.json()["page_url"])
    assert page.status_code == 200
    assert "Camille Martin" in page.text
    assert "GridLab" in page.text
    assert "Prévision de demande énergétique" in page.text
    assert "Python" in page.text
    assert "Describe the jobs you want" in page.text
    assert "Accepted contracts" in page.text
    assert "When an offer explicitly states another contract type" in page.text
    assert "Choose how far JobAuto may go" in page.text
    assert 'id="candidate-locale"' in page.text
    assert "Follow each offer's requested language" in page.text
    assert 'id="standard-work-authorization"' in page.text
    assert 'id="standard-availability"' in page.text
    assert 'id="standard-salary-expectation"' in page.text
    assert 'id="browser-preset"' in page.text
    assert "Advanced browser rules" in page.text
    assert "key=value" not in page.text
    assert "Largest readable one-page layout" not in page.text
    assert "Filled automatically from the CV" in page.text
    assert "<b>1</b> experiences" in page.text
    assert "<b>1</b> education entries" in page.text
    assert "École du numérique" in page.text
    assert "Adaptation" in page.text
    assert "Offer-dependent" in page.text
    assert "Derive a variation" in page.text
    assert "Save and open JobAuto" in page.text
    assert "localStorage.getItem(stageStorageKey)" in page.text
    assert "activeStage !== stages.length - 1" in page.text

    draft_url = f"/api/candidate-drafts/{status.json()['draft_id']}"
    draft = client.get(draft_url).json()
    update = {
        key: draft[key]
        for key in (
            "identity",
            "locale",
            "summary",
            "summary_source_block_ids",
            "letter_reference",
            "experiences",
            "skills",
            "projects",
            "education",
            "languages",
            "interests",
            "cv_layout",
            "project_lab",
            "search_preferences",
            "submission_preferences",
        )
    }
    update["expected_version"] = draft["version"]
    update["letter_reference"] = "Madame, Monsieur, je vous adresse ma candidature."
    update["skills"].append(
        {
            "name": "dbt",
            "category": "Data Engineering",
            "usage": "removable",
            "evidence": "transferable",
            "verification_warning": True,
            "source_block_ids": [],
        }
    )

    saved = client.put(draft_url, json=update)
    validated = client.post(f"{draft_url}/validate")

    assert saved.status_code == 200
    assert saved.json()["version"] == draft["version"] + 1
    assert saved.json()["skills"][-1]["name"] == "dbt"
    assert saved.json()["cv_layout"]["maximum_font_size_pt"] == 12.0
    assert validated.status_code == 200
    assert validated.json()["status"] == "validated"

    exported = client.post(f"{draft_url}/export")
    assert exported.status_code == 201
    assert exported.json()["candidate_id"].startswith("camille-martin-")
    latest_draft = client.get(draft_url).json()
    post_export_update = {**update, "expected_version": latest_draft["version"]}
    post_export_update["summary"] = f"{latest_draft['summary']} Updated after export."
    post_export_update["skills"] = latest_draft["skills"]
    updated_after_export = client.put(draft_url, json=post_export_update)
    assert updated_after_export.status_code == 200
    assert updated_after_export.json()["version"] > int(
        exported.json()["candidate_id"].rsplit("-v", maxsplit=1)[1]
    )
    workspace = client.get(exported.json()["page_url"])
    assert workspace.status_code == 200
    assert "Camille Martin" in workspace.text
    assert "Edit profile &amp; preferences" in workspace.text
    assert f'href="/candidate-drafts/{status.json()["draft_id"]}"' in workspace.text
    home = client.get("/")
    assert home.status_code == 200
    assert "Edit profile &amp; preferences" in home.text
    assert f'href="/candidate-drafts/{status.json()["draft_id"]}"' in home.text
    preview = client.get(exported.json()["preview_url"])
    assert preview.status_code == 200
    assert preview.content.startswith(b"%PDF")

    source_metadata = (
        tmp_path / "candidate-profiles" / exported.json()["candidate_id"] / "studio_source.json"
    )
    metadata = json.loads(source_metadata.read_text(encoding="utf-8"))
    source_metadata.write_text(
        json.dumps({key: metadata[key] for key in ("draft_id", "import_id")}),
        encoding="utf-8",
    )
    legacy_workspace = client.get(exported.json()["page_url"])
    assert "Edit profile &amp; preferences" in legacy_workspace.text

    source_metadata.write_text(
        json.dumps({**metadata, "candidate_id": "another-candidate"}),
        encoding="utf-8",
    )
    contaminated_workspace = client.get(exported.json()["page_url"])
    assert "Edit profile &amp; preferences" not in contaminated_workspace.text

    stale = client.put(draft_url, json=update)
    assert stale.status_code == 409
    assert "version conflict" in stale.json()["detail"]


def test_profile_workspace_shows_source_and_adaptation_contract(tmp_path: Path) -> None:
    response = _client(tmp_path).get("/profiles/alex-morgan")

    assert response.status_code == 200
    assert "Energy demand forecasting" in response.text
    assert "very faithful" in response.text.lower()
    assert "CV and preferences saved" in response.text
    assert "Find matching jobs" in response.text
    assert "Section adaptation rules" in response.text
    assert "Prepare my documents" in response.text
    assert "<select" not in response.text
    assert "policy-controls" not in response.text


def test_profile_preview_returns_real_pdf(tmp_path: Path) -> None:
    response = _client(tmp_path).get("/profiles/alex-morgan/preview.pdf")

    assert response.status_code == 200
    assert response.headers["content-type"] == "application/pdf"
    assert response.content.startswith(b"%PDF")


def test_unknown_profile_returns_404(tmp_path: Path) -> None:
    response = _client(tmp_path).get("/profiles/unknown")

    assert response.status_code == 404


def test_policy_section_post_is_not_reachable(tmp_path: Path) -> None:
    source_root = Path(__file__).resolve().parents[2]
    project_root = tmp_path / "project"
    profile_target = project_root / "config" / "profiles" / "example"
    profile_target.parent.mkdir(parents=True)
    shutil.copytree(source_root / "config" / "profiles" / "example", profile_target)
    client = TestClient(create_studio_app(project_root=project_root, state_root=tmp_path / "state"))
    policy_before = (profile_target / "adaptation_policy.yaml").read_text(encoding="utf-8")

    response = client.post(
        "/profiles/alex-morgan/policy/cv/summary",
        json={"fidelity": "locked", "target_lines": 3},
    )

    assert response.status_code == 404
    assert (profile_target / "adaptation_policy.yaml").read_text(encoding="utf-8") == policy_before


class FakeApplicationService:
    def __init__(self, root: Path) -> None:
        self.root = root
        self.record: RunRecord | None = None
        self.request = None

    def start(self, request) -> str:
        self.request = request
        run_id = "alex-morgan-studio-test"
        run_dir = self.root / run_id
        run_dir.mkdir(parents=True)
        snapshot = CandidateProfileRepository(request.profile_path.parent.parent).load_snapshot(
            request.profile_path
        )
        self.record = RunRecord(
            run_id=run_id,
            candidate_id="alex-morgan",
            profile_path=request.profile_path,
            status="pending",
            current_phase="pending",
            phase_history=["pending"],
            created_at="2026-07-16T10:00:00+00:00",
            updated_at="2026-07-16T10:00:00+00:00",
            offer_url=request.offer_url,
            offer_sha256="a" * 64,
            snapshot_hash=snapshot.snapshot_hash,
            context_hash="c" * 64,
            run_dir=run_dir,
        )
        (run_dir / "request.json").write_text(request.model_dump_json(indent=2), encoding="utf-8")
        return run_id

    def execute(self, _run_id: str) -> RunRecord:
        assert self.record is not None
        source_artifact_dir = self.record.run_dir / "source-artifacts"
        source_artifact_dir.mkdir()
        (source_artifact_dir / "cv.pdf").write_bytes(b"%PDF synthetic original cv")
        (self.record.run_dir / "application-brief.json").write_text(
            json.dumps(
                {
                    "company": "GridCo",
                    "role": "Data Engineer",
                    "role_family": "Data Engineering",
                    "language": "en",
                    "summary": "Build reliable operational data products.",
                    "responsibilities": ["Build reliable Python and SQL pipelines."],
                    "required_skills": ["Python", "SQL"],
                    "normalized_role": "Data Engineer",
                    "open_role": "Data Engineer",
                    "sector": "Energy",
                    "specialisations": ["Reliable data pipelines"],
                    "cv_angle": "Present verified pipeline ownership and data quality evidence.",
                    "letter_angle": "Connect reliable data delivery to GridCo operations.",
                    "adaptation_decisions": [
                        {
                            "surface": "both",
                            "decision": "Prioritise production-ready Python and SQL pipelines.",
                            "rationale": "This is the strongest evidence-backed match for the role.",
                            "fact_ids": ["experience.gridco.pipelines"],
                        }
                    ],
                    "requirements": [
                        {
                            "requirement_id": "req_python",
                            "requirement": "Build Python and SQL pipelines",
                            "source_excerpt": "build reliable Python and SQL pipelines",
                            "priority": "must",
                            "kind": "technical_skill",
                        }
                    ],
                    "evidence_mappings": [
                        {
                            "requirement_id": "req_python",
                            "evidence_level": "verified",
                            "fact_ids": ["experience.gridco.pipelines"],
                            "rationale": "The candidate has direct pipeline evidence.",
                        }
                    ],
                    "project_plan": {
                        "decision": "reframe",
                        "rationale": "Keep the strongest existing projects and change only their angle.",
                        "slots": [
                            {
                                "slot": 1,
                                "mode": "reframe",
                                "source_project_id": "reliable-pipelines",
                                "rationale": "Emphasise reliability and monitoring.",
                            }
                        ],
                    },
                    "skill_plan": {
                        "categories": ["Data Engineering"],
                        "items": [
                            {
                                "name": "Python",
                                "category": "Data Engineering",
                                "kind": "language",
                                "evidence_level": "verified",
                                "priority": "must",
                            },
                            {
                                "name": "SQL",
                                "category": "Data Engineering",
                                "kind": "language",
                                "evidence_level": "verified",
                                "priority": "must",
                            },
                        ],
                        "rationale": "Keep the visible competency section focused on the offer.",
                    },
                    "targeted_keywords": ["Python", "SQL", "data quality"],
                },
                indent=2,
            ),
            encoding="utf-8",
        )
        artifact_dir = self.record.run_dir / "artifacts"
        artifact_dir.mkdir()
        cv = artifact_dir / "cv.pdf"
        letter = artifact_dir / "letter.pdf"
        cv.write_bytes(b"%PDF synthetic cv")
        letter.write_bytes(b"%PDF synthetic letter")
        self.record = self.record.model_copy(
            update={
                "status": "completed",
                "current_phase": "completed",
                "phase_history": ["pending", "completed"],
                "artifacts": {
                    "cv": {
                        "pdf_path": str(cv),
                        "pdf_sha256": hashlib.sha256(cv.read_bytes()).hexdigest(),
                        "page_count": 1,
                    },
                    "letter": {
                        "pdf_path": str(letter),
                        "pdf_sha256": hashlib.sha256(letter.read_bytes()).hexdigest(),
                        "page_count": 1,
                    },
                },
                "review": {
                    "approved": True,
                    "score": 94,
                    "ats_score": 96,
                    "editorial_score": 93,
                    "adaptation_score": 92,
                    "blocking_issues": [],
                    "letter_argument": {
                        criterion: {
                            "state": "pass",
                            "rationale": "The rendered letter provides concrete support.",
                            "supporting_excerpt": "I am applying for the role.",
                        }
                        for criterion in (
                            "target_specificity",
                            "evidence_to_missions",
                            "candidate_contribution",
                            "motivation_credibility",
                            "tone_and_naturalness",
                        )
                    },
                    "requirement_coverage": [
                        {
                            "requirement_id": "req_python",
                            "coverage": "exact",
                            "rationale": "Python is supported by candidate evidence.",
                        }
                    ],
                },
            }
        )
        return self.record

    def get(self, _run_id: str) -> RunRecord:
        if self.record is None:
            raise FileNotFoundError
        return self.record


def test_studio_launches_the_injected_application_service(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[2]
    service = FakeApplicationService(tmp_path / "runs")
    client = TestClient(
        create_studio_app(
            project_root=project_root,
            state_root=tmp_path / "state",
            application_service=service,
        )
    )

    response = client.post(
        "/profiles/alex-morgan/runs",
        json={
            "company": "GridCo",
            "role": "Data Engineer",
            "offer_url": "https://example.test/job",
            "offer_text": "GridCo seeks a Data Engineer to build reliable Python and SQL pipelines.",
        },
    )

    assert response.status_code == 202
    assert response.json()["run_id"] == "alex-morgan-studio-test"
    assert service.request.company == "GridCo"
    assert service.request.max_repairs == 2
    status = client.get("/runs/alex-morgan-studio-test/status")
    assert status.json()["status"] == "completed"
    assert status.json()["adaptation_summary"]["project_plan"]["decision"] == "reframe"
    page = client.get("/runs/alex-morgan-studio-test")
    assert page.status_code == 200
    assert "Tracked application run" in page.text
    assert "GridCo" in page.text
    assert "Data Engineer" in page.text
    assert "Agent activity" in page.text
    assert "Understand offer" in page.text
    assert "Plan adaptation" in page.text
    assert "Write documents" in page.text
    assert "Verify PDFs" in page.text
    assert "Ready to apply" in page.text
    assert "Elapsed" in page.text
    assert "Codex tasks" in page.text
    assert "Codex executions" not in page.text
    assert "total_tokens_estimate" in page.text
    assert "ATS and document review" in page.text
    assert "94/100" in page.text
    assert "ATS 96" in page.text
    assert "Letter argument" in page.text
    assert "target specificity" in page.text
    assert "The rendered letter provides concrete support." in page.text
    assert "req_python" in page.text
    assert "Open CV PDF" in page.text
    assert "Open letter PDF" in page.text
    assert "Open original CV" in page.text
    assert "Compare the source and tailored CV" in page.text
    assert 'alt="Source CV first page"' in page.text
    assert 'alt="Tailored CV first page"' in page.text
    assert "What changed and why" in page.text
    assert "Planned adaptation strategy" in page.text
    assert "Present verified pipeline ownership" in page.text
    assert "Connect reliable data delivery" in page.text
    assert "Prioritise production-ready Python and SQL pipelines" in page.text
    assert "Project strategy" in page.text
    assert "Data Engineering" in page.text
    assert "Continue to application queue" in page.text
    original = client.get("/runs/alex-morgan-studio-test/artifacts/original-cv")
    assert original.status_code == 200
    assert original.headers["content-disposition"].startswith("inline;")
    assert original.content == b"%PDF synthetic original cv"

    attached = client.post("/runs/alex-morgan-studio-test/campaign")
    assert attached.status_code == 201
    assert attached.json()["page_url"].startswith("/campaigns/")
    campaign = client.get(attached.json()["page_url"])
    assert campaign.status_code == 200
    assert "Your application documents are ready" in campaign.text
    updated_run = client.get("/runs/alex-morgan-studio-test")
    assert f'href="{attached.json()["page_url"]}"' in updated_run.text


def test_studio_imports_an_observable_campaign_into_a_selected_tracker(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[2]
    service = FakeApplicationService(tmp_path / "runs")
    tracker = tmp_path / "tracker.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Postulations"
    for column, header in enumerate(TRACKER_COLUMNS, start=1):
        sheet.cell(1, column).value = header
    workbook.save(tracker)
    workbook.close()
    client = TestClient(
        create_studio_app(
            project_root=project_root,
            state_root=tmp_path / "state",
            application_service=service,
        )
    )
    description = (
        "Permanent Data Engineer role in Toulouse. Build reliable Python and SQL data "
        "pipelines on BigQuery, own data quality and production monitoring, collaborate "
        "with product teams, document technical choices and deliver measurable analytics "
        "outcomes. The team values testing, maintainability and operational ownership."
    )

    response = client.post(
        "/profiles/alex-morgan/campaigns",
        json={
            "tracker_path": str(tracker),
            "limit": 3,
            "candidates": [
                {
                    "company": "GridCo",
                    "role": "Data Engineer",
                    "url": "https://example.test/gridco-data",
                    "description": description,
                    "location": "Toulouse, France",
                    "contract_type": "permanent",
                    "posted_at": date.today().isoformat(),
                }
            ],
        },
    )

    assert response.status_code == 202
    assert response.json()["selected_count"] == 1
    status = client.get(response.json()["status_url"])
    assert status.status_code == 200
    assert status.json()["status"] == "completed"
    assert status.json()["completed_count"] == 1
    assert status.json()["blocked_count"] == 0
    assert status.json()["failed_count"] == 0
    resumed = client.post(f"/campaigns/{response.json()['campaign_id']}/resume")
    assert resumed.status_code == 202
    assert resumed.json()["campaign_id"] == response.json()["campaign_id"]
    page = client.get(response.json()["page_url"])
    assert page.status_code == 200
    assert "Application batch" in page.text
    assert "Your application documents are ready" in page.text
    assert "document packs ready" in page.text
    assert "Campaign finished" in page.text
    assert "CV and letter ready" in page.text
    assert "Apply automatically" in page.text
    assert "Fill everything, let me submit" in page.text
    assert "Simulation only" in page.text
    assert "window.location.reload()" not in page.text
    assert "GridCo" in page.text
    assert "selected" in page.text
    assert tracker.name in page.text
    assert str(tracker) not in page.text
    run_page = client.get("/runs/alex-morgan-studio-test")
    assert f'href="/campaigns/{response.json()["campaign_id"]}"' in run_page.text
    assert "← Application batch" in run_page.text

    submission = client.post(
        f"/campaigns/{response.json()['campaign_id']}/submission",
        json={"mode": "automatic"},
    )
    assert submission.status_code == 201
    assert submission.json()["status"] == "ready_for_chrome"
    assert submission.json()["ready_count"] == 1
    assert submission.json()["mode"] == "automatic"
    assert submission.json()["codex_url"].startswith("codex://new?prompt=")
    assert "queue mode is automatic" in submission.json()["codex_prompt"]
    submission_status = client.get(submission.json()["status_url"])
    assert submission_status.status_code == 200
    assert submission_status.json()["items"][0]["handoff_id"] is not None
    claimed = client.post(f"/campaigns/{response.json()['campaign_id']}/submission/claim-next")
    assert claimed.status_code == 200
    assert claimed.json()["packet"]["status"] == "claimed_for_chrome"
    assert claimed.json()["queue"]["claimed_count"] == 1
    second_claim = client.post(f"/campaigns/{response.json()['campaign_id']}/submission/claim-next")
    assert second_claim.status_code == 200
    assert second_claim.json()["packet"] is None
    queued_page = client.get(submission.json()["page_url"])
    assert "Apply with Codex Chrome" in queued_page.text
    assert 'id="submission-ready-count">0</span> ready for Chrome' in queued_page.text
    assert 'id="submission-progress-count">1</span> in progress' in queued_page.text
    assert "GridCo | Data Engineer" in queued_page.text

    handoff = client.post(
        f"/campaigns/{response.json()['campaign_id']}/handoffs/alex-morgan-studio-test"
    )
    assert handoff.status_code == 201
    assert handoff.json()["status"] == "claimed_for_chrome"
    handoff_page = client.get(handoff.json()["page_url"])
    assert handoff_page.status_code == 200
    assert "Codex Chrome Extension handoff" in handoff_page.text
    assert "Use the JobAuto Codex plugin" in handoff_page.text
    assert "Allow access to file URLs" in handoff_page.text
    assert "Rehashed after review" in handoff_page.text
    assert "Run sandbox Chrome proof" in handoff_page.text
    assert "Return to queue" in handoff_page.text
    assert "cv.pdf" in handoff_page.text
    assert "letter.pdf" in handoff_page.text
    assert str(tmp_path) not in handoff_page.text
    sandbox = client.get(f"/sandbox/apply/{handoff.json()['handoff_id']}")
    assert sandbox.status_code == 200
    assert "Submit application" in sandbox.text
    assert service.record is not None
    cv_path = Path(str(service.record.artifacts["cv"]["pdf_path"]))
    letter_path = Path(str(service.record.artifacts["letter"]["pdf_path"]))
    malformed = client.post(
        f"/sandbox/apply/{handoff.json()['handoff_id']}/submit",
        json={
            "full_name": "Alex Morgan",
            "email": "alex.morgan@example.test",
            "location": "Toulouse",
            "message": "Verified local test.",
            "cv": {"name": "cv.pdf", "content_base64": "not-base64!"},
            "letter": {"name": "letter.pdf", "content_base64": "not-base64!"},
        },
    )
    assert malformed.status_code == 422
    assert malformed.json()["detail"] == "cv is not valid base64"
    receipt = client.post(
        f"/sandbox/apply/{handoff.json()['handoff_id']}/submit",
        json={
            "full_name": "Alex Morgan",
            "email": "alex.morgan@example.test",
            "location": "Toulouse",
            "message": "I am interested in this verified sandbox role.",
            "cv": {
                "name": cv_path.name,
                "content_base64": base64.b64encode(cv_path.read_bytes()).decode("ascii"),
            },
            "letter": {
                "name": letter_path.name,
                "content_base64": base64.b64encode(letter_path.read_bytes()).decode("ascii"),
            },
        },
    )
    assert receipt.status_code == 201
    assert receipt.json()["status"] == "sandbox_verified"
    confirmation = client.get(receipt.json()["confirmation_url"])
    assert confirmation.status_code == 200
    assert "Application packet verified" in confirmation.text
    assert "No employer application was submitted" in confirmation.text
    verified_handoff = client.get(handoff.json()["page_url"])
    assert "Sandbox passed" in verified_handoff.text
    assert "Employer application submitted" not in verified_handoff.text
    workbook = load_workbook(tracker, read_only=True, data_only=True)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        assert sheet.cell(2, headers[TRACKER_COLUMNS[11]]).value is None
        assert sheet.cell(2, headers[TRACKER_COLUMNS[21]]).value is None
    finally:
        workbook.close()

    profile = client.get("/profiles/alex-morgan")
    assert "Find matching jobs" in profile.text
    assert "Recent campaigns" in profile.text
    assert response.json()["campaign_id"] in profile.text
    assert "Import structured offers manually" in profile.text
    assert "Offers JSON" in profile.text
    assert 'href="/profiles/alex-morgan/applications"' in profile.text

    dashboard = client.get("/profiles/alex-morgan/applications")
    assert dashboard.status_code == 200
    assert "Offers, documents and submissions" in dashboard.text
    assert "Role/profile fit estimate" in dashboard.text
    assert "Baseline and final CV ATS estimates" in dashboard.text
    assert "This is not an employer ATS score." in dashboard.text
    assert "GridCo" in dashboard.text
    assert "Compare CVs" in dashboard.text
    assert "data-status-detail" in dashboard.text
    assert "renderedIds.join('|') !== incomingIds.join('|')" in dashboard.text
    assert f'href="/runs/{service.record.run_id}"' in dashboard.text
    dashboard_status = client.get("/profiles/alex-morgan/applications/status")
    assert dashboard_status.status_code == 200
    assert dashboard_status.json()["summary"]["offers"] == 1
    assert dashboard_status.json()["summary"]["ready"] == 1
    assert dashboard_status.json()["items"][0]["final_ats"] == 96


def test_studio_discovery_handoff_feeds_the_existing_campaign_gate(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[2]
    service = FakeApplicationService(tmp_path / "runs")
    tracker = tmp_path / "tracker.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Postulations"
    for column, header in enumerate(TRACKER_COLUMNS, start=1):
        sheet.cell(1, column).value = header
    workbook.save(tracker)
    workbook.close()
    description = (
        "Permanent Data Engineer role in Toulouse. Build reliable Python and SQL data "
        "pipelines on BigQuery, own data quality and production monitoring, collaborate "
        "with product teams, document technical choices and deliver measurable analytics "
        "outcomes. The team values testing, maintainability and operational ownership."
    )

    class FakeDiscoveryAgent:
        def complete_json(self, _prompt, response_model, _phase):
            return response_model.model_validate(
                {
                    "offers": [
                        {
                            "company": "Fresh Grid",
                            "role": "Data Engineer",
                            "url": "https://example.test/fresh-grid",
                            "description": description,
                            "location": "Toulouse, France",
                            "contract_type": "permanent",
                            "posted_at": date.today().isoformat(),
                            "semantic_fit_score": 91,
                            "semantic_fit_rationale": (
                                "The role directly matches the candidate's configured data work."
                            ),
                        }
                    ],
                    "rejected_candidates": [],
                    "notes": "",
                    "END_OF_RESPONSE": True,
                }
            )

    client = TestClient(
        create_studio_app(
            project_root=project_root,
            state_root=tmp_path / "state",
            application_service=service,
            discovery_agent_factory=lambda _callback: FakeDiscoveryAgent(),
        )
    )

    prepared = client.post(
        "/profiles/alex-morgan/discoveries",
        json={
            "tracker_path": str(tracker),
            "requested_count": 3,
            "conversation_url": "https://chatgpt.com/c/jobauto-demo",
        },
    )
    assert prepared.status_code == 201
    page = client.get(prepared.json()["page_url"])
    assert page.status_code == 200
    assert "Finding your next 3 opportunities" in page.text
    assert "Searching current offers" in page.text
    assert "Where JobAuto is searching" in page.text
    assert "Codex web search" in page.text
    assert "official company career and ATS pages" in page.text
    assert "zero candidates during the search is normal" in page.text
    assert "Resume search" in page.text
    assert "automaticResumeAttempted" in page.text
    assert "Technical details" in page.text
    discovery = client.get(prepared.json()["status_url"]).json()
    assert discovery["status"] == "campaign_created"
    completed_discovery_page = client.get(prepared.json()["page_url"])
    assert "Open application batch" in completed_discovery_page.text
    assert f"/campaigns/{discovery['campaign_id']}" in completed_discovery_page.text
    campaign = client.get(f"/campaigns/{discovery['campaign_id']}/status")
    assert campaign.json()["status"] == "completed"
    assert campaign.json()["selected_count"] == 1
