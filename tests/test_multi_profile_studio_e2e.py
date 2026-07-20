from __future__ import annotations

from datetime import date
from pathlib import Path

from openpyxl import Workbook, load_workbook

from jobauto.application_service import RunApplicationService
from jobauto.candidate_profile import CvBackend
from jobauto.candidate_snapshot import CandidateProfileRepository
from jobauto.document_patch import CandidateDocumentDraft, CvAdaptationPatch, apply_cv_patch
from jobauto.excel_schema import CANDIDATE_ID_COLUMN, TRACKER_COLUMNS
from jobauto.models import (
    ApplicationBrief,
    CandidateApplicationReview,
    CandidateLetterDraft,
    LetterArgumentAssessment,
    LetterArgumentCriterionAssessment,
)
from jobauto.public_validation import extract_pdf_text
from jobauto.run_store import RunStore
from jobauto.studio_campaign import StudioCampaignService, StudioCampaignStore
from jobauto.submission_campaign import SubmissionCampaignService
from jobauto.submission_handoff import (
    SubmissionHandoffService,
    SubmissionHandoffStore,
    SubmissionReceipt,
)


def _passing_letter_argument() -> LetterArgumentAssessment:
    def criterion() -> LetterArgumentCriterionAssessment:
        return LetterArgumentCriterionAssessment(
            state="pass",
            rationale="The rendered letter provides concrete support for this criterion.",
            supporting_excerpt="I am applying for the",
        )

    return LetterArgumentAssessment(
        target_specificity=criterion(),
        evidence_to_missions=criterion(),
        candidate_contribution=criterion(),
        motivation_credibility=criterion(),
        tone_and_naturalness=criterion(),
    )


class CrossDomainDocumentPipeline:
    """Deterministic agent boundary; rendering and orchestration remain real."""

    def __init__(self, snapshot) -> None:
        self.snapshot = snapshot

    def generate_candidate_documents(self, row, _offer_text, *, project_lab_context=""):
        patch = CvAdaptationPatch()
        name = self.snapshot.cv_source.name
        letter = CandidateLetterDraft(
            greeting="Dear hiring team,",
            paragraphs=[
                f"I am applying for the {row.role} role at {row.company}.",
                "My verified experience and domain knowledge match the role's central responsibilities, and I would bring precise, collaborative delivery to the team.",
            ],
            closing=f"Kind regards,\n{name}",
            used_fact_ids=["identity.current"],
        ).validate_for_snapshot(self.snapshot)
        return CandidateDocumentDraft(
            brief=ApplicationBrief.model_construct(
                company=row.company,
                role=row.role,
                language="en",
                requirements=[],
            ),
            cv_patch=patch,
            cv=apply_cv_patch(self.snapshot, patch),
            letter=letter,
        )

    def review_candidate_documents(
        self,
        _row,
        _package,
        cv_rendered,
        letter_rendered,
        _offer_text,
    ):
        assert cv_rendered.page_count == letter_rendered.page_count == 1
        assert len(cv_rendered.pdf_sha256) == len(letter_rendered.pdf_sha256) == 64
        return CandidateApplicationReview(
            approved=True,
            score=90,
            ats_score=90,
            editorial_score=90,
            adaptation_score=90,
            blocking_issues=[],
            warnings=[],
            letter_argument=_passing_letter_argument(),
            requirement_coverage=[],
        )

    def repair_candidate_documents(self, *_args, **_kwargs):
        raise AssertionError("The cross-domain happy path must not require repair")


def _tracker(path: Path) -> Path:
    workbook = Workbook()
    try:
        sheet = workbook.active
        sheet.title = "Postulations"
        for column, header in enumerate(TRACKER_COLUMNS, start=1):
            sheet.cell(1, column).value = header
        workbook.save(path)
    finally:
        workbook.close()
    return path


def _candidate(
    *, company: str, role: str, url: str, description: str, location: str
) -> dict[str, object]:
    return {
        "company": company,
        "role": role,
        "url": url,
        "description": description,
        "location": location,
        "contract_type": "permanent",
        "posted_at": "2026-07-17",
    }


def test_two_profiles_share_one_tracker_without_cross_contamination_and_restart(
    tmp_path: Path,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profiles = project_root / "tests" / "fixtures" / "cross_domain_profiles"
    offers = project_root / "tests" / "fixtures" / "offers"
    repository = CandidateProfileRepository(profiles)
    application = RunApplicationService(
        repository=repository,
        store=RunStore(tmp_path / "runs"),
        pipeline_factory=lambda snapshot, _context: CrossDomainDocumentPipeline(snapshot),
    )
    campaign_store = StudioCampaignStore(tmp_path / "campaigns")
    campaigns = StudioCampaignService(
        repository=repository,
        application_service=application,
        store=campaign_store,
    )
    tracker = _tracker(tmp_path / "applications.xlsx")

    frontend = campaigns.create(
        profile_path=profiles / "frontend" / "profile.yaml",
        tracker_path=tracker,
        candidates=[
            _candidate(
                company="Interface Studio",
                role="Frontend Software Engineer",
                url="https://example.test/jobs/frontend",
                description=(offers / "frontend_software_engineer_en.txt").read_text(
                    encoding="utf-8"
                ),
                location="Bordeaux, France",
            )
        ],
        limit=1,
        today=date(2026, 7, 18),
    )
    regulatory = campaigns.create(
        profile_path=profiles / "regulatory" / "profile.yaml",
        tracker_path=tracker,
        candidates=[
            _candidate(
                company="Medica Europe",
                role="Regulatory Affairs Specialist",
                url="https://example.test/jobs/regulatory",
                description=(offers / "regulatory_affairs_specialist_en.txt").read_text(
                    encoding="utf-8"
                ),
                location="Lyon, France",
            )
        ],
        limit=1,
        today=date(2026, 7, 18),
    )
    frontend = campaigns.execute(frontend.campaign_id)
    regulatory = campaigns.execute(regulatory.campaign_id)

    assert frontend.status == regulatory.status == "completed"
    assert frontend.items[0].tracker_artifacts_synced is True
    assert regulatory.items[0].tracker_artifacts_synced is True
    assert frontend.items[0].run_id != regulatory.items[0].run_id
    assert (
        application.get(frontend.items[0].run_id).context_hash
        != application.get(regulatory.items[0].run_id).context_hash
    )
    for record in (
        application.get(frontend.items[0].run_id),
        application.get(regulatory.items[0].run_id),
    ):
        for kind in ("cv", "letter"):
            artifact = record.artifacts[kind]
            pdf_path = Path(str(artifact["pdf_path"]))
            assert pdf_path.read_bytes().startswith(b"%PDF-")
            assert artifact["page_count"] == 1
            assert int(artifact["extracted_text_characters"]) > 200
        cv_metrics = record.artifacts["cv"]["layout_metrics"]
        assert 10.0 <= float(cv_metrics["font_size_pt"]) <= 12.0
        assert 1.10 <= float(cv_metrics["line_height_ratio"]) <= 1.50
        assert float(cv_metrics["vertical_coverage_ratio"]) >= 0.82
        assert cv_metrics["requires_density_review"] is False
    regulatory_snapshot = repository.load_snapshot(profiles / "regulatory" / "profile.yaml")
    assert regulatory_snapshot.profile.cv_backend is CvBackend.SOURCE_PRESERVING
    regulatory_cv = Path(
        str(application.get(regulatory.items[0].run_id).artifacts["cv"]["pdf_path"])
    )
    regulatory_text = extract_pdf_text(regulatory_cv)
    assert all(
        section in regulatory_text
        for section in ("Certifications", "Professional Memberships", "Volunteering")
    )

    handoff_store = SubmissionHandoffStore(tmp_path / "handoffs")
    handoffs = SubmissionHandoffService(
        repository=repository,
        campaign_service=campaigns,
        run_reader=application,
        store=handoff_store,
    )
    submissions = SubmissionCampaignService(
        campaign_service=campaigns,
        handoff_service=handoffs,
    )
    prepared_frontend = submissions.prepare(frontend.campaign_id)
    prepared_regulatory = submissions.prepare(regulatory.campaign_id)

    assert prepared_frontend.candidate_id == "noah-williams"
    assert prepared_regulatory.candidate_id == "sofia-martin"
    for prepared in (prepared_frontend, prepared_regulatory):
        evidence = tmp_path / f"{prepared.candidate_id}-confirmation.png"
        evidence.write_bytes(b"confirmation")
        handoffs.record_receipt(
            prepared.items[0].handoff_id,
            SubmissionReceipt(
                status="submitted",
                portal="sandbox",
                evidence_path=evidence,
                uploaded_files=["cv", "letter"],
            ),
        )

    restarted_campaigns = StudioCampaignService(
        repository=repository,
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )
    restarted_handoffs = SubmissionHandoffService(
        repository=repository,
        campaign_service=restarted_campaigns,
        run_reader=application,
        store=SubmissionHandoffStore(tmp_path / "handoffs"),
    )
    restarted_submissions = SubmissionCampaignService(
        campaign_service=restarted_campaigns,
        handoff_service=restarted_handoffs,
    )
    assert restarted_submissions.get(frontend.campaign_id).status == "submitted"
    assert restarted_submissions.get(regulatory.campaign_id).status == "submitted"

    workbook = load_workbook(tracker, read_only=False, data_only=True)
    try:
        sheet = workbook["Postulations"]
        columns = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        rows = {
            sheet.cell(row, columns[CANDIDATE_ID_COLUMN]).value: row
            for row in range(2, sheet.max_row + 1)
        }
        assert set(rows) == {"noah-williams", "sofia-martin"}
        expected_urls = {
            "noah-williams": "https://example.test/jobs/frontend",
            "sofia-martin": "https://example.test/jobs/regulatory",
        }
        for candidate_id, row in rows.items():
            assert sheet.cell(row, columns["Lien offre"]).value == expected_urls[candidate_id]
            assert sheet.cell(row, columns["CV PDF"]).hyperlink is not None
            assert sheet.cell(row, columns["Lettre PDF"]).hyperlink is not None
            assert sheet.cell(row, columns[TRACKER_COLUMNS[11]]).value == "Oui"
            assert sheet.cell(row, columns[TRACKER_COLUMNS[21]]).value == "submitted"
    finally:
        workbook.close()
