from __future__ import annotations

from concurrent.futures import ThreadPoolExecutor
from datetime import date
from pathlib import Path
from types import SimpleNamespace

from openpyxl import Workbook, load_workbook

from jobauto.candidate_snapshot import CandidateProfileRepository
from jobauto.excel_schema import CANDIDATE_ID_COLUMN, TRACKER_COLUMNS
from jobauto.offer_catalog import OfferCandidate
from jobauto.run_store import RunRecord
from jobauto.studio_campaign import (
    StudioCampaignRecord,
    StudioCampaignService,
    StudioCampaignStore,
    _parse_description_experience_years,
    _parse_experience_years,
    _run_observability,
    _to_search_offer,
)


def test_experience_parser_uses_year_context_instead_of_degree_numbers() -> None:
    assert _parse_experience_years("0-3 years of data engineering experience") == 3
    assert _parse_experience_years("Au moins 2 ans d'expérience") == 2
    assert _parse_experience_years("5+ years; Bac+5 degree") == 5
    assert (
        _parse_experience_years(
            "Expérience requise sans minimum chiffré; diplôme Bac+5 en informatique"
        )
        is None
    )


def test_description_experience_parser_reads_candidate_requirements_only() -> None:
    assert (
        _parse_description_experience_years(
            "Our company was founded 20 years ago. Candidates must have at least "
            "5 years of regulatory affairs experience."
        )
        == 5
    )
    assert (
        _parse_description_experience_years(
            "Our team brings 20 years of experience to customers. The role is open now."
        )
        is None
    )


def test_description_experience_parser_keeps_requirement_heading_context() -> None:
    description = """CANDIDATE REQUIREMENTS
Technical Experience: 5-6+ years of experience building production pipelines.
Adaptability & Ownership: Work independently in a fast-moving environment.
WHY JOIN US
Our team brings 20 years of experience to customers.
"""

    assert _parse_description_experience_years(description) == 6


def test_run_observability_counts_terminal_calls_once() -> None:
    run = type(
        "ObservedRun",
        (),
        {
            "agent_events": [
                {"call_id": "a", "phase": "cv_writer", "status": "running"},
                {
                    "call_id": "a",
                    "phase": "cv_writer",
                    "status": "succeeded",
                    "latency_ms": 1200,
                    "total_tokens_estimate": 2500,
                },
                {
                    "call_id": "b",
                    "phase": "repair",
                    "status": "succeeded",
                    "latency_ms": 800,
                    "total_tokens_estimate": 500,
                },
            ]
        },
    )()

    assert _run_observability(run) == (2, 1, 2000, 3000, "repair", "succeeded")


def test_campaign_offer_fallback_rejects_experience_gap_hidden_in_description() -> None:
    project_root = Path(__file__).resolve().parents[1]
    snapshot = CandidateProfileRepository(project_root / "config" / "profiles").load_snapshot(
        project_root / "config" / "profiles" / "example" / "profile.yaml"
    )
    candidate = OfferCandidate(
        company="Medica Europe",
        role="Regulatory Affairs Specialist",
        url="https://example.test/jobs/regulatory-affairs-specialist",
        description=(
            "The role supports technical documentation and regulatory submissions. "
            "Candidates must have at least 5 years of regulatory affairs experience. "
            "The team works across quality, clinical and manufacturing functions."
        ),
    )

    offer = _to_search_offer(candidate, today=date(2026, 7, 18))
    evaluation = snapshot.search_preferences.evaluate(offer, today=date(2026, 7, 18))

    assert offer.experience_years == 5
    assert evaluation.eligible is False
    assert any(item.criterion == "experience_years" for item in evaluation.blockers)


def test_demo_campaign_keeps_an_experience_stretch_and_prepares_only_best_offer(
    tmp_path: Path,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    application = FakeApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
        demo_fast=True,
    )
    record = service.create(
        profile_path=profile_path,
        tracker_path=_empty_tracker(tmp_path / "tracker.xlsx"),
        candidates=[
            {
                "company": "Best Stretch",
                "role": "Data Engineer",
                "url": "https://example.test/jobs/best-stretch",
                "description": _full_description("Requires 5 years of experience."),
                "experience_required": "5 years",
                "semantic_fit_score": 95,
            },
            {
                "company": "Reserve",
                "role": "Analytics Engineer",
                "url": "https://example.test/jobs/reserve",
                "description": _full_description("Open to early-career candidates."),
                "experience_required": "0-2 years",
                "semantic_fit_score": 90,
            },
        ],
        limit=2,
    )

    assert record.demo_fast is True
    assert record.selected_count == 1
    assert record.reserve_count == 1
    assert record.items[0].decision == "selected"
    assert "ranked stretch" in " ".join(record.items[0].reasons)
    assert application.requests[0].demo_fast is True
    assert application.requests[0].max_repairs == 1


class FakeApplicationService:
    def __init__(self, root: Path) -> None:
        self.root = root
        self.requests = []
        self.records: dict[str, RunRecord] = {}

    def start(self, request) -> str:
        self.requests.append(request)
        run_id = f"run-{len(self.requests):08d}"
        run_dir = self.root / run_id
        run_dir.mkdir(parents=True)
        self.records[run_id] = RunRecord(
            run_id=run_id,
            candidate_id="alex-morgan",
            profile_path=request.profile_path,
            status="pending",
            current_phase="pending",
            phase_history=["pending"],
            created_at="2026-07-17T10:00:00+00:00",
            updated_at="2026-07-17T10:00:00+00:00",
            offer_url=request.offer_url,
            offer_sha256="a" * 64,
            snapshot_hash="b" * 64,
            context_hash="c" * 64,
            run_dir=run_dir,
        )
        return run_id

    def execute(self, run_id: str) -> RunRecord:
        artifact_dir = self.records[run_id].run_dir / "artifacts"
        artifact_dir.mkdir()
        cv = artifact_dir / "cv.pdf"
        letter = artifact_dir / "letter.pdf"
        cv.write_bytes(b"%PDF synthetic cv")
        letter.write_bytes(b"%PDF synthetic letter")
        record = self.records[run_id].model_copy(
            update={
                "status": "completed",
                "current_phase": "completed",
                "artifacts": {
                    "cv": {"pdf_path": str(cv)},
                    "letter": {"pdf_path": str(letter)},
                },
            }
        )
        self.records[run_id] = record
        return record

    def get(self, run_id: str) -> RunRecord:
        return self.records[run_id]


class BlockingFirstApplicationService(FakeApplicationService):
    def __init__(self, root: Path) -> None:
        super().__init__(root)
        self.executions: list[str] = []

    def execute(self, run_id: str) -> RunRecord:
        self.executions.append(run_id)
        if run_id == "run-00000001":
            record = self.records[run_id].model_copy(
                update={
                    "status": "blocked",
                    "current_phase": "blocked",
                    "blockers": ["terminal candidate fit gap: req_critical"],
                }
            )
            self.records[run_id] = record
            return record
        return super().execute(run_id)


class AlwaysBlockingApplicationService(FakeApplicationService):
    def __init__(self, root: Path) -> None:
        super().__init__(root)
        self.executions: list[str] = []

    def execute(self, run_id: str) -> RunRecord:
        self.executions.append(run_id)
        record = self.records[run_id].model_copy(
            update={
                "status": "blocked",
                "current_phase": "blocked",
                "blockers": ["terminal candidate fit gap: req_critical"],
            }
        )
        self.records[run_id] = record
        return record


class CancellingAfterFirstApplicationService(FakeApplicationService):
    campaign_service: StudioCampaignService | None = None
    campaign_id: str | None = None

    def execute(self, run_id: str) -> RunRecord:
        record = super().execute(run_id)
        if run_id == "run-00000001" and self.campaign_service and self.campaign_id:
            self.campaign_service.request_cancel(self.campaign_id)
        return record


def _tracker(path: Path, duplicate_url: str) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Postulations"
    for column, header in enumerate(TRACKER_COLUMNS, start=1):
        sheet.cell(1, column).value = header
    sheet.cell(2, 1).value = "Existing Co"
    sheet.cell(2, 2).value = "Existing role"
    sheet.cell(2, 3).value = duplicate_url
    workbook.save(path)
    workbook.close()
    return path


def _empty_tracker(path: Path) -> Path:
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Postulations"
    for column, header in enumerate(TRACKER_COLUMNS, start=1):
        sheet.cell(1, column).value = header
    workbook.save(path)
    workbook.close()
    return path


def _full_description(marker: str) -> str:
    return (
        f"{marker} permanent Data Engineer role in Toulouse. "
        "Build reliable Python and SQL data pipelines on BigQuery, own data quality, "
        "work with product teams, document operational decisions, and deliver measurable "
        "analytics outcomes. The team values maintainability, collaboration, testing, "
        "clear communication, and practical production ownership."
    )


def test_campaign_deduplicates_filters_appends_and_starts_generic_runs(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _tracker(tmp_path / "tracker.xlsx", "https://example.test/existing?utm_source=x")
    application = FakeApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )

    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": "Existing Co",
                "role": "Data Engineer",
                "url": "https://example.test/existing",
                "description": _full_description("Duplicate"),
                "contract_type": "permanent",
            },
            {
                "company": "Grid Systems",
                "role": "Data Engineer",
                "url": "https://example.test/jobs/data-engineer?utm_campaign=batch",
                "description": _full_description("Fresh"),
                "location": "Toulouse, France",
                "contract_type": "permanent",
                "experience_required": "0-3 years",
                "posted_at": "2026-07-15",
            },
            {
                "company": "Thin Summary",
                "role": "Analytics Engineer",
                "url": "https://example.test/jobs/thin",
                "description": "Python and SQL role.",
            },
        ],
        limit=5,
        today=date(2026, 7, 17),
    )

    assert [item.decision for item in record.items] == ["duplicate", "selected", "rejected"]
    selected = record.items[1]
    assert selected.evaluation is not None
    assert selected.evaluation.eligible is True
    assert selected.excel_row == 3
    assert selected.run_id == "run-00000001"
    assert application.requests[0].offer_text == _full_description("Fresh")

    workbook = load_workbook(tracker, read_only=False, data_only=True)
    try:
        sheet = workbook["Postulations"]
        assert sheet.cell(3, 1).value == "Grid Systems"
        assert (
            sheet.cell(3, 3).value == "https://example.test/jobs/data-engineer?utm_campaign=batch"
        )
        assert sheet.cell(3, 6).value == _full_description("Fresh")
    finally:
        workbook.close()

    completed = service.execute(record.campaign_id)
    assert completed.status == "completed"
    assert completed.items[1].run_status == "completed"
    assert completed.items[1].tracker_artifacts_synced is True
    assert service.get(record.campaign_id).items[1].run_phase == "completed"
    workbook = load_workbook(tracker, read_only=False, data_only=True)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        assert sheet.cell(3, headers["CV PDF"]).hyperlink is not None
        assert sheet.cell(3, headers["Lettre PDF"]).hyperlink is not None
    finally:
        workbook.close()


def test_completed_direct_run_enters_campaign_without_regeneration(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _empty_tracker(tmp_path / "tracker.xlsx")
    application = FakeApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )
    description = _full_description("Direct")
    run_id = application.start(
        SimpleNamespace(
            profile_path=profile_path,
            offer_url="https://example.test/jobs/direct",
        )
    )
    application.execute(run_id)

    campaign = service.attach_completed_run(
        profile_path=profile_path,
        tracker_path=tracker,
        offer=OfferCandidate(
            company="Direct Co",
            role="Data Engineer",
            url="https://example.test/jobs/direct",
            description=description,
            location="Toulouse, France",
            contract_type="permanent",
        ),
        run_id=run_id,
    )

    assert campaign.status == "completed"
    assert campaign.items[0].run_id == run_id
    assert campaign.items[0].tracker_artifacts_synced is True
    assert len(application.requests) == 1
    assert (
        service.attach_completed_run(
            profile_path=profile_path,
            tracker_path=tracker,
            offer=campaign.items[0].offer,
            run_id=run_id,
        ).campaign_id
        == campaign.campaign_id
    )
    assert len(application.requests) == 1

    workbook = load_workbook(tracker, read_only=False, data_only=True)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        assert sheet.cell(2, headers["Lien offre"]).value == "https://example.test/jobs/direct"
        assert sheet.cell(2, headers["CV PDF"]).hyperlink is not None
        assert sheet.cell(2, headers["Lettre PDF"]).hyperlink is not None
    finally:
        workbook.close()


def test_campaign_stops_between_applications_and_can_resume(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    application = CancellingAfterFirstApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )
    record = service.create(
        profile_path=profile_path,
        tracker_path=_empty_tracker(tmp_path / "tracker.xlsx"),
        candidates=[
            {
                "company": "First Legal Tech",
                "role": "Data Engineer",
                "url": "https://example.test/jobs/first",
                "description": _full_description("First"),
                "location": "Toulouse, France",
                "contract_type": "permanent",
                "experience_required": "0-3 years",
                "posted_at": "2026-07-15",
            },
            {
                "company": "Second Legal Tech",
                "role": "Data Engineer",
                "url": "https://example.test/jobs/second",
                "description": _full_description("Second"),
                "location": "Toulouse, France",
                "contract_type": "permanent",
                "experience_required": "0-3 years",
                "posted_at": "2026-07-15",
            },
        ],
        limit=2,
        today=date(2026, 7, 17),
    )
    application.campaign_service = service
    application.campaign_id = record.campaign_id

    cancelled = service.execute(record.campaign_id)

    assert cancelled.status == "cancelled"
    assert cancelled.items[0].run_status == "completed"
    assert cancelled.items[1].run_status == "pending"

    service.resume(record.campaign_id)
    completed = service.execute(record.campaign_id)

    assert completed.status == "completed"
    assert all(
        item.run_status == "completed" for item in completed.items if item.decision == "selected"
    )


def test_campaign_resume_requeues_a_terminal_run_without_duplicating_tracker_rows(
    tmp_path: Path,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _empty_tracker(tmp_path / "tracker.xlsx")
    application = FakeApplicationService(tmp_path / "runs")
    store = StudioCampaignStore(tmp_path / "campaigns")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=store,
    )
    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": "Grid Systems",
                "role": "Data Engineer",
                "url": "https://example.test/jobs/data-engineer",
                "description": _full_description("Retry"),
                "location": "Toulouse, France",
                "contract_type": "permanent",
            }
        ],
        limit=1,
        today=date(2026, 7, 17),
    )
    item = record.items[0]
    original_run_id = item.run_id
    original_excel_row = item.excel_row
    item.run_status = "failed"
    item.run_phase = "failed"
    item.run_blockers = ["old failure"]
    store.save(record.model_copy(update={"status": "partial", "items": record.items}))

    resumed = service.resume(record.campaign_id)

    retried = resumed.items[0]
    assert retried.run_id != original_run_id
    assert retried.run_status == "pending"
    assert retried.run_phase == "pending"
    assert retried.run_blockers == []
    assert retried.excel_row == original_excel_row
    assert len(application.requests) == 2
    workbook = load_workbook(tracker, read_only=True)
    try:
        assert workbook["Postulations"].max_row == 2
    finally:
        workbook.close()

    completed = service.execute(record.campaign_id)
    assert completed.status == "completed"
    assert completed.items[0].run_status == "completed"


def test_campaign_resume_can_recover_a_running_record_after_process_restart(
    tmp_path: Path,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _empty_tracker(tmp_path / "tracker.xlsx")
    application = FakeApplicationService(tmp_path / "runs")
    store = StudioCampaignStore(tmp_path / "campaigns")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=store,
    )
    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": "Grid Systems",
                "role": "Data Engineer",
                "url": "https://example.test/jobs/data-engineer",
                "description": _full_description("Interrupted"),
                "location": "Toulouse, France",
                "contract_type": "permanent",
            }
        ],
        limit=1,
        today=date(2026, 7, 17),
    )
    item = record.items[0]
    old_run_id = item.run_id
    item.run_status = "running"
    item.run_phase = "agent:cv_writer:running"
    store.save(record.model_copy(update={"status": "running", "items": record.items}))

    resumed = service.resume(record.campaign_id, allow_interrupted=True)

    assert resumed.status == "ready"
    assert resumed.items[0].run_id != old_run_id
    assert resumed.items[0].run_status == "pending"
    assert resumed.items[0].excel_row == item.excel_row


def test_campaign_prefers_semantic_role_fit_over_incidental_keyword_overlap(
    tmp_path: Path,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = (
        project_root
        / "tests"
        / "fixtures"
        / "cross_domain_profiles"
        / "regulatory"
        / "profile.yaml"
    )
    tracker = _empty_tracker(tmp_path / "tracker.xlsx")
    application = FakeApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(profile_path.parent.parent),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )

    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": "Campus Operations",
                "role": "Regulatory Building Operations Coordinator",
                "url": "https://example.test/jobs/building-operations",
                "description": (
                    "Coordinate permanent building operations in Lyon, France, including "
                    "vendor access, facilities documentation, safety inspections and service "
                    "schedules. The role communicates with regulated medical-device tenants "
                    "and records EU MDR and ISO 13485 references supplied by those tenants, "
                    "but owns neither regulatory submissions nor product quality evidence."
                ),
                "location": "Lyon, France",
                "contract_type": "permanent",
                "semantic_fit_score": 18,
                "semantic_fit_rationale": (
                    "The central function is facilities coordination rather than regulatory affairs."
                ),
            },
            {
                "company": "Health Products",
                "role": "Quality and Regulatory Affairs Specialist",
                "url": "https://example.test/jobs/quality-regulatory",
                "description": (
                    "Join a permanent Lyon team preparing technical files, coordinating "
                    "product submissions, maintaining quality-system evidence and answering "
                    "authority questions. Work cross-functionally with clinical, engineering "
                    "and manufacturing colleagues to keep regulated products compliant from "
                    "development through market surveillance."
                ),
                "location": "Lyon, France",
                "contract_type": "permanent",
                "semantic_fit_score": 94,
                "semantic_fit_rationale": (
                    "The central responsibilities match the candidate's regulatory evidence."
                ),
            },
        ],
        limit=1,
        today=date(2026, 7, 18),
    )

    assert [item.decision for item in record.items] == ["not_selected", "selected"]
    assert application.requests[0].role == "Quality and Regulatory Affairs Specialist"


def test_completed_run_does_not_attach_artifacts_to_another_candidates_row(
    tmp_path: Path,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _empty_tracker(tmp_path / "tracker.xlsx")
    application = FakeApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )
    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": "Grid Systems",
                "role": "Data Engineer",
                "url": "https://example.test/jobs/data-engineer",
                "description": _full_description("Fresh"),
                "location": "Toulouse, France",
                "contract_type": "permanent",
                "posted_at": "2026-07-15",
            }
        ],
        limit=1,
        today=date(2026, 7, 17),
    )
    selected = record.items[0]
    assert selected.excel_row is not None
    workbook = load_workbook(tracker)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        sheet.cell(selected.excel_row, headers[CANDIDATE_ID_COLUMN]).value = "another-candidate"
        workbook.save(tracker)
    finally:
        workbook.close()

    completed = service.execute(record.campaign_id)

    item = completed.items[0]
    assert item.run_status == "completed"
    assert item.tracker_artifacts_synced is False
    assert item.tracker_sync_error is not None
    assert "belongs to another candidate" in item.tracker_sync_error
    workbook = load_workbook(tracker)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        sheet.cell(selected.excel_row, headers[CANDIDATE_ID_COLUMN]).value = "alex-morgan"
        workbook.save(tracker)
    finally:
        workbook.close()

    retried = service.execute(record.campaign_id)

    retried_item = retried.items[0]
    assert retried_item.tracker_artifacts_synced is True
    assert retried_item.tracker_sync_error is None
    workbook = load_workbook(tracker, read_only=False, data_only=True)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        assert sheet.cell(selected.excel_row, headers["CV PDF"]).hyperlink is not None
        assert sheet.cell(selected.excel_row, headers["Lettre PDF"]).hyperlink is not None
    finally:
        workbook.close()


def test_campaign_limit_keeps_eligible_offers_observable(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _tracker(tmp_path / "tracker.xlsx", "https://example.test/existing")
    application = FakeApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )

    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": company,
                "role": "Data Engineer",
                "url": f"https://example.test/{company.casefold()}",
                "description": _full_description(company),
                "location": "Toulouse",
                "contract_type": "permanent",
            }
            for company in ("First", "Second")
        ],
        limit=1,
        today=date(2026, 7, 17),
    )

    assert sum(item.decision == "selected" for item in record.items) == 1
    assert sum(item.decision == "not_selected" for item in record.items) == 1
    assert len(application.requests) == 1

    expanded = service.expand(record.campaign_id)

    assert expanded.selected_count == 2
    assert expanded.reserve_count == 0
    assert expanded.processing_limit == 2
    assert len(application.requests) == 2
    assert expanded.items[1].reasons == ["Promoted from the eligible reserve by the candidate."]
    completed = service.execute(record.campaign_id)
    assert completed.status == "completed"
    assert sum(item.run_status == "completed" for item in completed.items) == 2


def test_campaign_rechecks_reserve_availability_before_generation(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _tracker(tmp_path / "tracker.xlsx", "https://example.test/existing")
    application = FakeApplicationService(tmp_path / "runs")

    class FakeAvailabilityVerifier:
        def verify(self, url: str):
            if url.endswith("/closed"):
                return SimpleNamespace(status="unavailable", reason="Offer page is closed.")
            return SimpleNamespace(status="available", reason="Offer page is open.")

    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
        availability_verifier=FakeAvailabilityVerifier(),
    )
    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": company,
                "role": "Data Engineer",
                "url": f"https://example.test/{slug}",
                "description": _full_description(company),
                "location": "Toulouse",
                "contract_type": "permanent",
            }
            for company, slug in (("First", "first"), ("Closed", "closed"), ("Open", "open"))
        ],
        limit=1,
        today=date(2026, 7, 17),
    )

    expanded = service.expand(record.campaign_id, additional_count=2)

    closed = next(item for item in expanded.items if item.offer.company == "Closed")
    opened = next(item for item in expanded.items if item.offer.company == "Open")
    assert closed.decision == "rejected"
    assert closed.reasons == ["Offer unavailable before generation: Offer page is closed."]
    assert opened.decision == "selected"
    assert len(application.requests) == 2


def test_campaign_respects_candidate_processing_limit(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _tracker(tmp_path / "tracker.xlsx", "https://example.test/existing")
    application = FakeApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )

    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": f"Company {index}",
                "role": "Data Engineer",
                "url": f"https://example.test/company-{index}",
                "description": _full_description(f"Company {index}"),
                "location": "Toulouse",
                "contract_type": "permanent",
            }
            for index in range(8)
        ],
        limit=100,
        today=date(2026, 7, 17),
    )

    assert record.requested_limit == 100
    assert record.processing_limit == 5
    assert record.effective_limit == 5
    assert record.selected_count == 5
    assert len(application.requests) == 5
    assert sum(item.decision == "not_selected" for item in record.items) == 3
    assert record.items[-1].reasons == [
        "Eligible but outside the candidate's configured campaign limit."
    ]


def test_campaign_promotes_next_ranked_offer_when_first_run_is_blocked(
    tmp_path: Path,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _tracker(tmp_path / "tracker.xlsx", "https://example.test/existing")
    application = BlockingFirstApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )
    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": company,
                "role": "Data Engineer",
                "url": f"https://example.test/{company.casefold()}",
                "description": _full_description(company),
                "location": "Toulouse",
                "contract_type": "permanent",
            }
            for company in ("First", "Second", "Third")
        ],
        limit=1,
        today=date(2026, 7, 17),
    )

    completed = service.execute(record.campaign_id)

    assert completed.status == "completed"
    assert application.executions == ["run-00000001", "run-00000002"]
    assert len(application.requests) == 2
    assert completed.items[0].run_status == "blocked"
    assert completed.items[1].run_status == "completed"
    assert completed.items[1].reasons == [
        "Promoted after an earlier candidate did not produce documents."
    ]
    assert completed.items[2].decision == "not_selected"


def test_demo_campaign_stops_after_one_fallback_instead_of_draining_reserves(
    tmp_path: Path,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    tracker = _tracker(tmp_path / "tracker.xlsx", "https://example.test/existing")
    application = AlwaysBlockingApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(project_root / "config" / "profiles"),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
        demo_fast=True,
    )
    record = service.create(
        profile_path=profile_path,
        tracker_path=tracker,
        candidates=[
            {
                "company": company,
                "role": "Data Engineer",
                "url": f"https://example.test/{company.casefold()}",
                "description": _full_description(company),
                "location": "Toulouse",
                "contract_type": "permanent",
            }
            for company in ("First", "Second", "Third", "Fourth")
        ],
        limit=3,
        today=date(2026, 7, 17),
    )

    completed = service.execute(record.campaign_id)

    assert completed.status == "partial"
    assert application.executions == ["run-00000001", "run-00000002"]
    assert len(application.requests) == 2
    assert completed.items[0].run_status == "blocked"
    assert completed.items[1].run_status == "blocked"
    assert all(item.run_id is None for item in completed.items[2:])


def test_campaign_ranking_comes_from_the_selected_candidate_profile(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    repository = CandidateProfileRepository(project_root / "config" / "profiles")
    candidates = [
        {
            "company": "Applied Models",
            "role": "Machine Learning Engineer",
            "url": "https://example.test/ml-engineer",
            "description": (
                "Permanent Machine Learning Engineer role in Lyon. Build and evaluate "
                "production image-classification models with Python, PyTorch and MLflow. "
                "Own reproducible experiments, error analysis, model monitoring, technical "
                "documentation and collaboration with product teams. The role values practical "
                "deployment, reliable evaluation and clear communication across disciplines."
            ),
            "location": "Lyon, France",
            "contract_type": "permanent",
        }
    ]

    def create_for(profile: str, name: str):
        application = FakeApplicationService(tmp_path / f"runs-{name}")
        service = StudioCampaignService(
            repository=repository,
            application_service=application,
            store=StudioCampaignStore(tmp_path / f"campaigns-{name}"),
        )
        tracker = _tracker(tmp_path / f"tracker-{name}.xlsx", "https://example.test/existing")
        return service.create(
            profile_path=project_root / "config" / "profiles" / profile / "profile.yaml",
            tracker_path=tracker,
            candidates=candidates,
            limit=1,
            today=date(2026, 7, 17),
        )

    alex = create_for("example", "alex")
    jamie = create_for("example-b", "jamie")

    assert alex.items[0].evaluation is not None
    assert jamie.items[0].evaluation is not None
    assert jamie.items[0].evaluation.score > alex.items[0].evaluation.score
    assert jamie.candidate_id == "jamie-chen"
    assert alex.candidate_id == "alex-morgan"


def test_shared_tracker_deduplicates_per_candidate_not_globally(tmp_path: Path) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profiles = project_root / "tests" / "fixtures" / "cross_domain_profiles"
    offers = project_root / "tests" / "fixtures" / "offers"
    tracker = _empty_tracker(tmp_path / "shared-tracker.xlsx")
    application = FakeApplicationService(tmp_path / "runs")
    service = StudioCampaignService(
        repository=CandidateProfileRepository(profiles),
        application_service=application,
        store=StudioCampaignStore(tmp_path / "campaigns"),
    )
    shared_url = "https://example.test/jobs/shared-role"

    frontend = service.create(
        profile_path=profiles / "frontend" / "profile.yaml",
        tracker_path=tracker,
        candidates=[
            {
                "company": "Interface Studio",
                "role": "Frontend Software Engineer",
                "url": shared_url,
                "description": (offers / "frontend_software_engineer_en.txt").read_text(
                    encoding="utf-8"
                ),
                "location": "Bordeaux, France",
                "contract_type": "permanent",
                "posted_at": "2026-07-17",
            }
        ],
        limit=1,
        today=date(2026, 7, 18),
    )
    regulatory = service.create(
        profile_path=profiles / "regulatory" / "profile.yaml",
        tracker_path=tracker,
        candidates=[
            {
                "company": "Medica Europe",
                "role": "Regulatory Affairs Specialist",
                "url": shared_url,
                "description": (offers / "regulatory_affairs_specialist_en.txt").read_text(
                    encoding="utf-8"
                ),
                "location": "Lyon, France",
                "contract_type": "permanent",
                "posted_at": "2026-07-17",
            }
        ],
        limit=1,
        today=date(2026, 7, 18),
    )

    assert frontend.items[0].decision == "selected"
    assert regulatory.items[0].decision == "selected"

    repeated = service.create(
        profile_path=profiles / "frontend" / "profile.yaml",
        tracker_path=tracker,
        candidates=[frontend.items[0].offer.model_dump(mode="json")],
        limit=1,
        today=date(2026, 7, 18),
    )
    assert repeated.items[0].decision == "duplicate"

    workbook = load_workbook(tracker, read_only=False, data_only=True)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        owner_column = headers[CANDIDATE_ID_COLUMN]
        assert sheet.max_row == 3
        assert [sheet.cell(row, owner_column).value for row in (2, 3)] == [
            "noah-williams",
            "sofia-martin",
        ]
        assert sheet.column_dimensions[sheet.cell(1, owner_column).column_letter].hidden is True
    finally:
        workbook.close()


def test_concurrent_campaign_creation_does_not_duplicate_one_candidate_url(
    tmp_path: Path,
) -> None:
    project_root = Path(__file__).resolve().parents[1]
    profiles = project_root / "tests" / "fixtures" / "cross_domain_profiles"
    offer_text = (
        project_root / "tests" / "fixtures" / "offers" / "frontend_software_engineer_en.txt"
    ).read_text(encoding="utf-8")
    tracker = _empty_tracker(tmp_path / "concurrent-tracker.xlsx")
    candidate = {
        "company": "Interface Studio",
        "role": "Frontend Software Engineer",
        "url": "https://example.test/jobs/concurrent-role",
        "description": offer_text,
        "location": "Bordeaux, France",
        "contract_type": "permanent",
        "posted_at": "2026-07-17",
    }

    def create(index: int) -> StudioCampaignRecord:
        service = StudioCampaignService(
            repository=CandidateProfileRepository(profiles),
            application_service=FakeApplicationService(tmp_path / f"runs-{index}"),
            store=StudioCampaignStore(tmp_path / f"campaigns-{index}"),
        )
        return service.create(
            profile_path=profiles / "frontend" / "profile.yaml",
            tracker_path=tracker,
            candidates=[candidate],
            limit=1,
            today=date(2026, 7, 18),
        )

    with ThreadPoolExecutor(max_workers=2) as executor:
        records = list(executor.map(create, range(2)))

    assert sorted(record.items[0].decision for record in records) == ["duplicate", "selected"]
    workbook = load_workbook(tracker, read_only=True, data_only=True)
    try:
        assert workbook["Postulations"].max_row == 2
    finally:
        workbook.close()


def test_campaign_store_lists_recent_records_for_one_candidate(tmp_path: Path) -> None:
    store = StudioCampaignStore(tmp_path / "campaigns")
    common = {
        "profile_path": tmp_path / "profile.yaml",
        "tracker_path": tmp_path / "tracker.xlsx",
        "status": "ready",
        "created_at": "2026-07-17T10:00:00+00:00",
        "requested_limit": 3,
        "items": [],
    }
    for campaign_id, candidate_id, updated_at in (
        ("campaign-alex-old", "alex-morgan", "2026-07-17T10:00:00+00:00"),
        ("campaign-jamie", "jamie-chen", "2026-07-17T11:00:00+00:00"),
        ("campaign-alex-new", "alex-morgan", "2026-07-17T12:00:00+00:00"),
    ):
        store.create(
            StudioCampaignRecord(
                campaign_id=campaign_id,
                candidate_id=candidate_id,
                updated_at=updated_at,
                campaign_dir=store.root / campaign_id,
                **common,
            )
        )

    assert [record.campaign_id for record in store.list_for_candidate("alex-morgan")] == [
        "campaign-alex-new",
        "campaign-alex-old",
    ]
