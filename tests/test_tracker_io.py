from __future__ import annotations

import os
from pathlib import Path

from openpyxl import Workbook, load_workbook

from jobauto import tracker_io


def test_atomic_workbook_save_retries_transient_windows_lock(tmp_path: Path, monkeypatch) -> None:
    target = tmp_path / "applications.xlsx"
    workbook = Workbook()
    workbook.active["A1"] = "ready"
    original_replace = os.replace
    calls = 0

    def transient_replace(source: str | Path, destination: str | Path) -> None:
        nonlocal calls
        calls += 1
        if calls < 3:
            raise PermissionError("transient file lock")
        original_replace(source, destination)

    monkeypatch.setattr(tracker_io.os, "replace", transient_replace)
    monkeypatch.setattr(tracker_io.time, "sleep", lambda _seconds: None)

    try:
        tracker_io.save_workbook_atomically(workbook, target)
    finally:
        workbook.close()

    assert calls == 3
    loaded = load_workbook(target, read_only=True, data_only=True)
    try:
        assert loaded.active["A1"].value == "ready"
    finally:
        loaded.close()
    assert list(tmp_path.glob(".applications.*.xlsx")) == []
