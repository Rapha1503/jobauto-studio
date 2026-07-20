from __future__ import annotations

import hashlib
from concurrent.futures import ThreadPoolExecutor
from dataclasses import replace
from pathlib import Path

import pytest
from openpyxl import Workbook, load_workbook

from jobauto.candidate_snapshot import CandidateProfileRepository
from jobauto.discovery_handoff import OfferAvailabilityResult
from jobauto.excel_schema import CANDIDATE_ID_COLUMN, TRACKER_COLUMNS
from jobauto.run_store import RunRecord
from jobauto.studio_campaign import StudioCampaignItem, StudioCampaignRecord
from jobauto.submission_campaign import SubmissionCampaignService
from jobauto.submission_handoff import (
    SubmissionHandoffService,
    SubmissionHandoffStore,
    SubmissionReceipt,
)
from jobauto.submission_preferences import SubmissionMode


class FakeCampaignService:
    def __init__(self, record: StudioCampaignRecord) -> None:
        self.record = record

    def get(self, _campaign_id: str) -> StudioCampaignRecord:
        return self.record


class FakeRunReader:
    def __init__(self, record: RunRecord) -> None:
        self.record = record

    def get(self, _run_id: str) -> RunRecord:
        return self.record


class MappingRunReader:
    def __init__(self, records: dict[str, RunRecord]) -> None:
        self.records = records

    def get(self, run_id: str) -> RunRecord:
        return self.records[run_id]


class ClosedOfferVerifier:
    def verify(self, url: str) -> OfferAvailabilityResult:
        return OfferAvailabilityResult(
            url=url,
            status="unavailable",
            status_code=200,
            reason="Offer page states that the position is closed.",
        )


def _sha(path: Path) -> str:
    return hashlib.sha256(path.read_bytes()).hexdigest()


def _fixture(tmp_path: Path):
    project_root = Path(__file__).resolve().parents[1]
    profile_path = project_root / "config" / "profiles" / "example" / "profile.yaml"
    repository = CandidateProfileRepository(project_root / "config" / "profiles")
    snapshot = repository.load_snapshot(profile_path)
    tracker = tmp_path / "tracker.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Postulations"
    for column, header in enumerate(TRACKER_COLUMNS, 1):
        sheet.cell(1, column).value = header
    sheet.cell(2, 1).value = "GridCo"
    sheet.cell(2, 2).value = "Data Engineer"
    sheet.cell(2, 3).value = "https://example.test/grid"
    workbook.save(tracker)
    workbook.close()
    run_dir = tmp_path / "run"
    run_dir.mkdir()
    cv = run_dir / "cv.pdf"
    letter = run_dir / "letter.pdf"
    cv.write_bytes(b"%PDF verified cv")
    letter.write_bytes(b"%PDF verified letter")
    run = RunRecord(
        run_id="alex-morgan-run-0001",
        candidate_id="alex-morgan",
        profile_path=profile_path,
        status="completed",
        current_phase="completed",
        phase_history=["completed"],
        created_at="2026-07-17T10:00:00+00:00",
        updated_at="2026-07-17T10:00:00+00:00",
        offer_url="https://example.test/grid",
        offer_sha256="a" * 64,
        snapshot_hash=snapshot.snapshot_hash,
        context_hash="c" * 64,
        run_dir=run_dir,
        artifacts={
            "cv": {"pdf_path": str(cv), "pdf_sha256": _sha(cv), "page_count": 1},
            "letter": {
                "pdf_path": str(letter),
                "pdf_sha256": _sha(letter),
                "page_count": 1,
            },
        },
        review={"approved": True},
    )
    campaign = StudioCampaignRecord(
        campaign_id="alex-morgan-campaign-0001",
        candidate_id="alex-morgan",
        profile_path=profile_path,
        tracker_path=tracker,
        status="completed",
        created_at="2026-07-17T10:00:00+00:00",
        updated_at="2026-07-17T10:00:00+00:00",
        requested_limit=1,
        campaign_dir=tmp_path / "campaign",
        items=[
            StudioCampaignItem(
                offer={
                    "company": "GridCo",
                    "role": "Data Engineer",
                    "url": "https://example.test/grid",
                    "description": "Complete offer text " * 20,
                },
                canonical_url="https://example.test/grid",
                decision="selected",
                excel_row=2,
                run_id=run.run_id,
                run_status="completed",
            )
        ],
    )
    service = SubmissionHandoffService(
        repository=repository,
        campaign_service=FakeCampaignService(campaign),
        run_reader=FakeRunReader(run),
        store=SubmissionHandoffStore(tmp_path / "handoffs"),
    )
    return service, campaign, run, cv


def test_handoff_rehashes_approved_artifacts_and_receipt_updates_tracker(tmp_path: Path) -> None:
    service, campaign, _run, _cv = _fixture(tmp_path)

    handoff = service.prepare(campaign.campaign_id, "alex-morgan-run-0001")

    assert handoff.status == "ready_for_chrome"
    assert handoff.candidate_identity is not None
    assert handoff.candidate_identity.email == "alex.morgan@example.test"
    assert handoff.candidate_form_profile is not None
    assert handoff.candidate_form_profile.experiences[0].organization == "Northwind Energy"
    assert handoff.candidate_form_profile.experiences[0].role == "Data Engineer"
    assert "Python" in handoff.candidate_form_profile.experiences[0].description[0]
    assert {artifact.kind for artifact in handoff.artifacts} == {"cv", "letter"}
    assert handoff.preferences.max_applications_per_campaign == 5

    service.store.save(handoff.model_copy(update={"candidate_form_profile": None}))
    backfilled = service.prepare(campaign.campaign_id, "alex-morgan-run-0001")
    assert backfilled.candidate_form_profile is not None
    assert backfilled.candidate_form_profile.experiences[0].organization == "Northwind Energy"

    evidence = tmp_path / "confirmation.png"
    evidence.write_bytes(b"png")
    submitted = service.record_receipt(
        handoff.handoff_id,
        SubmissionReceipt(
            status="submitted",
            portal="greenhouse",
            evidence_path=evidence,
            uploaded_files=["cv", "letter"],
        ),
    )

    assert submitted.status == "submitted"
    assert submitted.tracker_sync_error is None
    workbook = load_workbook(campaign.tracker_path, read_only=False, data_only=True)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        assert sheet.cell(2, headers[TRACKER_COLUMNS[11]]).value == "Oui"
        assert sheet.cell(2, headers[TRACKER_COLUMNS[21]]).value == "submitted"
        assert sheet.cell(2, headers[TRACKER_COLUMNS[24]]).hyperlink is not None
        assert sheet.cell(2, headers[CANDIDATE_ID_COLUMN]).value == "alex-morgan"
    finally:
        workbook.close()


def test_chrome_claim_is_exclusive_and_can_be_released(tmp_path: Path) -> None:
    service, campaign, _run, _cv = _fixture(tmp_path)
    prepared = service.prepare(campaign.campaign_id, "alex-morgan-run-0001")

    claimed = service.claim_next(campaign.campaign_id)

    assert claimed is not None
    assert claimed.handoff_id == prepared.handoff_id
    assert claimed.status == "claimed_for_chrome"
    assert service.claim_next(campaign.campaign_id) is None
    released = service.release_claim(prepared.handoff_id)
    assert released.status == "ready_for_chrome"
    assert service.claim_next(campaign.campaign_id) is not None


def test_receipt_does_not_mutate_tracker_row_owned_by_another_candidate(
    tmp_path: Path,
) -> None:
    service, campaign, _run, _cv = _fixture(tmp_path)
    handoff = service.prepare(campaign.campaign_id, "alex-morgan-run-0001")
    workbook = load_workbook(campaign.tracker_path)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        sheet.cell(2, headers[CANDIDATE_ID_COLUMN]).value = "another-candidate"
        workbook.save(campaign.tracker_path)
    finally:
        workbook.close()
    evidence = tmp_path / "confirmation.png"
    evidence.write_bytes(b"png")
    receipt = SubmissionReceipt(
        status="submitted",
        portal="greenhouse",
        evidence_path=evidence,
        uploaded_files=["cv", "letter"],
    )

    submitted = service.record_receipt(
        handoff.handoff_id,
        receipt,
    )

    assert submitted.status == "submitted"
    assert submitted.tracker_sync_error is not None
    assert "belongs to another candidate" in submitted.tracker_sync_error
    workbook = load_workbook(campaign.tracker_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        assert sheet.cell(2, headers[CANDIDATE_ID_COLUMN]).value == "another-candidate"
        assert sheet.cell(2, headers[TRACKER_COLUMNS[11]]).value is None
        assert sheet.cell(2, headers[TRACKER_COLUMNS[21]]).value is None
    finally:
        workbook.close()

    workbook = load_workbook(campaign.tracker_path)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        sheet.cell(2, headers[CANDIDATE_ID_COLUMN]).value = "alex-morgan"
        workbook.save(campaign.tracker_path)
    finally:
        workbook.close()

    retried = service.record_receipt(handoff.handoff_id, receipt)

    assert retried.tracker_sync_error is None
    workbook = load_workbook(campaign.tracker_path, read_only=True, data_only=True)
    try:
        sheet = workbook["Postulations"]
        headers = {
            str(sheet.cell(1, column).value): column for column in range(1, sheet.max_column + 1)
        }
        assert sheet.cell(2, headers[TRACKER_COLUMNS[11]]).value == "Oui"
        assert sheet.cell(2, headers[TRACKER_COLUMNS[21]]).value == "submitted"
    finally:
        workbook.close()


def test_handoff_blocks_when_artifact_hash_changed(tmp_path: Path) -> None:
    service, campaign, _run, cv = _fixture(tmp_path)
    cv.write_bytes(b"%PDF changed after review")

    handoff = service.prepare(campaign.campaign_id, "alex-morgan-run-0001")

    assert handoff.status == "blocked"
    assert "cv_hash_mismatch" in handoff.blockers


def test_receipt_replay_is_idempotent_and_submitted_state_is_terminal(
    tmp_path: Path,
) -> None:
    service, campaign, _run, _cv = _fixture(tmp_path)
    handoff = service.prepare(campaign.campaign_id, "alex-morgan-run-0001")
    evidence = tmp_path / "confirmation.png"
    evidence.write_bytes(b"png")
    receipt = SubmissionReceipt(
        status="submitted",
        portal="greenhouse",
        submitted_at="2026-07-18T10:00:00+00:00",
        evidence_path=evidence,
        uploaded_files=["cv", "letter"],
    )

    first = service.record_receipt(handoff.handoff_id, receipt)
    replay = service.record_receipt(handoff.handoff_id, receipt)

    assert replay == first
    with pytest.raises(ValueError, match="conflicting receipt"):
        service.record_receipt(
            handoff.handoff_id,
            receipt.model_copy(update={"portal": "lever"}),
        )


def test_dry_run_rejects_employer_submission_receipt(tmp_path: Path) -> None:
    service, campaign, _run, _cv = _fixture(tmp_path)
    handoff = service.prepare(campaign.campaign_id, "alex-morgan-run-0001")
    handoff = handoff.model_copy(
        update={
            "status": "dry_run",
            "preferences": handoff.preferences.model_copy(update={"mode": SubmissionMode.DRY_RUN}),
        }
    )
    service.store.save(handoff)
    evidence = tmp_path / "confirmation.png"
    evidence.write_bytes(b"png")

    with pytest.raises(ValueError, match="dry-run handoff"):
        service.record_receipt(
            handoff.handoff_id,
            SubmissionReceipt(
                status="submitted",
                portal="greenhouse",
                evidence_path=evidence,
            ),
        )


def test_handoff_preparation_is_idempotent_across_service_restarts(tmp_path: Path) -> None:
    service, campaign, run, _cv = _fixture(tmp_path)

    first = service.prepare(campaign.campaign_id, run.run_id)
    restarted = SubmissionHandoffService(
        repository=service.repository,
        campaign_service=service.campaign_service,
        run_reader=service.run_reader,
        store=SubmissionHandoffStore(tmp_path / "handoffs"),
    )
    second = restarted.prepare(campaign.campaign_id, run.run_id)

    assert second.handoff_id == first.handoff_id
    assert len(restarted.store.list_for_campaign(campaign.campaign_id)) == 1


def test_handoff_preparation_is_idempotent_under_concurrent_requests(tmp_path: Path) -> None:
    service, campaign, run, _cv = _fixture(tmp_path)

    with ThreadPoolExecutor(max_workers=2) as executor:
        records = list(
            executor.map(
                lambda _index: service.prepare(campaign.campaign_id, run.run_id),
                range(2),
            )
        )

    assert records[0].handoff_id == records[1].handoff_id
    assert len(service.store.list_for_campaign(campaign.campaign_id)) == 1


def test_submission_campaign_prepares_queue_and_reflects_terminal_receipt(
    tmp_path: Path,
) -> None:
    handoffs, campaign, run, _cv = _fixture(tmp_path)
    service = SubmissionCampaignService(
        campaign_service=handoffs.campaign_service,
        handoff_service=handoffs,
    )

    prepared = service.prepare(campaign.campaign_id)
    resumed = service.prepare(campaign.campaign_id)

    assert prepared.status == "ready_for_chrome"
    assert prepared.ready_count == 1
    assert resumed.items[0].handoff_id == prepared.items[0].handoff_id
    assert len(handoffs.store.list_for_campaign(campaign.campaign_id)) == 1

    evidence = tmp_path / "confirmation.png"
    evidence.write_bytes(b"png")
    handoffs.record_receipt(
        prepared.items[0].handoff_id,
        SubmissionReceipt(
            status="submitted",
            portal="greenhouse",
            evidence_path=evidence,
            uploaded_files=["cv", "letter"],
        ),
    )

    completed = service.get(campaign.campaign_id)
    assert completed.status == "submitted"
    assert completed.submitted_count == 1
    assert completed.ready_count == 0
    assert completed.items[0].run_id == run.run_id


def test_submission_campaign_blocks_offer_closed_after_document_generation(
    tmp_path: Path,
) -> None:
    handoffs, campaign, _run, _cv = _fixture(tmp_path)
    service = SubmissionCampaignService(
        campaign_service=handoffs.campaign_service,
        handoff_service=handoffs,
        availability_verifier=ClosedOfferVerifier(),
    )

    summary = service.prepare(campaign.campaign_id)

    assert summary.status == "blocked"
    assert summary.blocked_count == 1
    assert summary.ready_count == 0
    assert summary.items[0].blockers[0].startswith("offer_unavailable_before_submission:")


def test_submission_campaign_keeps_unfinished_run_visible(tmp_path: Path) -> None:
    handoffs, campaign, _run, _cv = _fixture(tmp_path)
    campaign.items[0].run_status = "running"
    service = SubmissionCampaignService(
        campaign_service=handoffs.campaign_service,
        handoff_service=handoffs,
    )

    summary = service.prepare(campaign.campaign_id)

    assert summary.status == "waiting_for_documents"
    assert summary.waiting_count == 1
    assert summary.items[0].status == "waiting_for_documents"
    assert handoffs.store.list_for_campaign(campaign.campaign_id) == []


def test_submission_campaign_classifies_failed_generation_as_blocked(tmp_path: Path) -> None:
    handoffs, campaign, _run, _cv = _fixture(tmp_path)
    campaign.items[0].run_status = "failed"
    campaign.items[0].run_blockers = ["render_failed"]
    service = SubmissionCampaignService(
        campaign_service=handoffs.campaign_service,
        handoff_service=handoffs,
    )

    summary = service.prepare(campaign.campaign_id)

    assert summary.status == "blocked"
    assert summary.blocked_count == 1
    assert summary.items[0].blockers == ["render_failed"]


def test_submission_campaign_handles_mixed_batch_and_restart_without_duplicates(
    tmp_path: Path,
) -> None:
    handoffs, campaign, first_run, _cv = _fixture(tmp_path)
    blocked_run = first_run.model_copy(
        update={
            "run_id": "alex-morgan-run-0002",
            "review": {"approved": False},
        }
    )
    waiting_run = first_run.model_copy(
        update={
            "run_id": "alex-morgan-run-0003",
            "status": "running",
            "current_phase": "generating_documents",
            "artifacts": {},
            "review": None,
        }
    )
    for run, company in ((blocked_run, "BlockedCo"), (waiting_run, "WaitingCo")):
        campaign.items.append(
            StudioCampaignItem(
                offer={
                    "company": company,
                    "role": "Data Engineer",
                    "url": f"https://example.test/{company.casefold()}",
                    "description": "Complete offer text " * 20,
                },
                canonical_url=f"https://example.test/{company.casefold()}",
                decision="selected",
                excel_row=len(campaign.items) + 2,
                run_id=run.run_id,
                run_status=run.status,
                run_blockers=list(run.blockers),
            )
        )
    reader = MappingRunReader({run.run_id: run for run in (first_run, blocked_run, waiting_run)})
    handoffs = SubmissionHandoffService(
        repository=handoffs.repository,
        campaign_service=handoffs.campaign_service,
        run_reader=reader,
        store=SubmissionHandoffStore(tmp_path / "handoffs"),
    )
    service = SubmissionCampaignService(
        campaign_service=handoffs.campaign_service,
        handoff_service=handoffs,
    )

    first = service.prepare(campaign.campaign_id)
    restarted_handoffs = SubmissionHandoffService(
        repository=handoffs.repository,
        campaign_service=handoffs.campaign_service,
        run_reader=reader,
        store=SubmissionHandoffStore(tmp_path / "handoffs"),
    )
    restarted = SubmissionCampaignService(
        campaign_service=restarted_handoffs.campaign_service,
        handoff_service=restarted_handoffs,
    ).prepare(campaign.campaign_id)

    assert first.status == "partial"
    assert (first.ready_count, first.blocked_count, first.waiting_count) == (1, 1, 1)
    assert [item.status for item in first.items] == [
        "ready_for_chrome",
        "blocked",
        "waiting_for_documents",
    ]
    assert [item.handoff_id for item in restarted.items] == [
        item.handoff_id for item in first.items
    ]
    assert len(restarted_handoffs.store.list_for_campaign(campaign.campaign_id)) == 2


def test_submission_campaign_mode_is_frozen_by_first_persisted_handoff(
    tmp_path: Path,
    monkeypatch: pytest.MonkeyPatch,
) -> None:
    handoffs, campaign, _run, _cv = _fixture(tmp_path)
    service = SubmissionCampaignService(
        campaign_service=handoffs.campaign_service,
        handoff_service=handoffs,
    )
    prepared = service.prepare(campaign.campaign_id)
    original_load = handoffs.repository.load_snapshot
    original_snapshot = original_load(campaign.profile_path)
    changed_snapshot = replace(
        original_snapshot,
        _submission_preferences=original_snapshot.submission_preferences.model_copy(
            update={"mode": SubmissionMode.AUTOMATIC}
        ),
    )
    monkeypatch.setattr(handoffs.repository, "load_snapshot", lambda _path: changed_snapshot)

    refreshed = service.get(campaign.campaign_id)

    assert prepared.mode is SubmissionMode.CONFIRM
    assert refreshed.mode is SubmissionMode.CONFIRM


def test_submission_campaign_can_choose_automatic_or_confirm_before_chrome_claim(
    tmp_path: Path,
) -> None:
    handoffs, campaign, _run, _cv = _fixture(tmp_path)
    service = SubmissionCampaignService(
        campaign_service=handoffs.campaign_service,
        handoff_service=handoffs,
    )

    automatic = service.prepare(campaign.campaign_id, mode=SubmissionMode.AUTOMATIC)
    assert automatic.mode is SubmissionMode.AUTOMATIC
    assert automatic.ready_count == 1
    persisted = handoffs.store.list_for_campaign(campaign.campaign_id)[0]
    assert persisted.preferences.mode is SubmissionMode.AUTOMATIC

    confirm = service.prepare(campaign.campaign_id, mode=SubmissionMode.CONFIRM)
    assert confirm.mode is SubmissionMode.CONFIRM
    assert handoffs.store.list_for_campaign(campaign.campaign_id)[0].preferences.mode is (
        SubmissionMode.CONFIRM
    )


def test_submission_mode_cannot_change_after_chrome_claim(tmp_path: Path) -> None:
    handoffs, campaign, _run, _cv = _fixture(tmp_path)
    service = SubmissionCampaignService(
        campaign_service=handoffs.campaign_service,
        handoff_service=handoffs,
    )
    service.prepare(campaign.campaign_id, mode=SubmissionMode.CONFIRM)
    handoffs.claim_next(campaign.campaign_id)

    with pytest.raises(ValueError, match="cannot change"):
        service.prepare(campaign.campaign_id, mode=SubmissionMode.AUTOMATIC)
